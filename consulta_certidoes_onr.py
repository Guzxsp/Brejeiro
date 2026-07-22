# -*- coding: utf-8 -*-
"""
╔══════════════════════════════════════════════════════════════╗
║            CONSULTA CERTIDÕES ONR — v2.1                    ║
║  Verifica protocolos em arquivos locais e no OneDrive.      ║
║  Python 3.9+ · Tkinter · openpyxl (aba planilha)           ║
║  v2.1 — Sistema de abas + Atualização automática           ║
╚══════════════════════════════════════════════════════════════╝
"""

from __future__ import annotations

import sys

# ══════════════════════════════════════════════════════════════
#  VERSÃO E CONFIGURAÇÃO DE ATUALIZAÇÃO
# ══════════════════════════════════════════════════════════════
APP_VERSION = "2.1.0"
UPDATE_CHECK_URL = "https://arrozbrejeiro.sharepoint.com/..."  # URL do arquivo de versão
UPDATE_EXE_URL = "https://arrozbrejeiro.sharepoint.com/..."    # URL do EXE atualizado
UPDATE_FOLDER = r"C:\Users\ANPSOJA3\OneDrive - BREJEIRO\Atualizacoes_CertidoesONR"

import glob as _glob
import json
import os
import shutil
import subprocess
import tempfile
import threading
import tkinter as tk
import urllib.request
import urllib.error
import zipfile
from tkinter import filedialog, messagebox, ttk


# ══════════════════════════════════════════════════════════════
#  PALETA DE CORES
# ══════════════════════════════════════════════════════════════
C_BG          = "#F5F7FA"   # fundo geral
C_HEADER      = "#1B3A5C"   # cabeçalho (azul escuro)
C_CARD        = "#FFFFFF"   # fundo dos cartões
C_ACCENT      = "#1565C0"   # azul principal
C_SUCCESS     = "#2E7D32"   # verde (encontrado local)
C_DRIVE       = "#0078D4"   # azul OneDrive
C_DANGER      = "#C62828"   # vermelho (não encontrado)
C_MUTED       = "#607D8B"   # cinza (botão limpar)
C_PRINT       = "#6A1B9A"   # roxo (botão imprimir)
C_COUNTER_BG  = "#E3F2FD"   # fundo do contador
C_STATUS_BG   = "#ECEFF1"   # barra de status
C_EXCEL       = "#217346"   # verde Excel (aba planilha)
C_EXCEL_DARK  = "#185C38"   # verde Excel escuro (hover)

# ══════════════════════════════════════════════════════════════
#  PASTA PADRÃO DO ONEDRIVE (editável)
# ══════════════════════════════════════════════════════════════
ONEDRIVE_DEFAULT = (
    r"C:\Users\ANPSOJA3\OneDrive - BREJEIRO\Comercial Soja - ANP-"
    r"Controle de Certidões de Penhor (Goiás) - Controle de Certidões"
    r" de Penhor (Goiás)\Safra 2025-2026\Certidões de Forma Digital"
)

ONEDRIVE_WEB_URL = (
    "https://arrozbrejeiro.sharepoint.com/:f:/s/ANP-COMPRASSOJA-"
    "ControledeCertidesdePenhorGois/IgD00w0ApkavRpdUJPCryk4uAVmr1JRw"
    "tRIbOpes3MZMpUA?e=gF8r91"
)

# Subpath da planilha de controle dentro do OneDrive
PLANILHA_SUBPATH = (
    r"OneDrive - BREJEIRO\Comercial Soja - ANP-Controle de Certidões de Penhor (Goiás)"
    r" - Controle de Certidões de Penhor (Goiás)\Safra 2025-2026\Controle de Certidões 2026.xlsx"
)
PLANILHA_NOME = "Controle de Certidões 2026.xlsx"


# ══════════════════════════════════════════════════════════════
#  UTILITÁRIO: localizar planilha dinamicamente
# ══════════════════════════════════════════════════════════════
def _find_planilha() -> str | None:
    """Itera C:\\Users\\*\\ procurando a planilha pelo subpath."""
    users_dir = r"C:\Users"
    try:
        for user in os.listdir(users_dir):
            candidate = os.path.join(users_dir, user, PLANILHA_SUBPATH)
            if os.path.isfile(candidate):
                return candidate
    except Exception:
        pass
    return None


# ══════════════════════════════════════════════════════════════
#  HELPERS DE WIDGET (funções livres reutilizáveis)
# ══════════════════════════════════════════════════════════════
def _make_btn(
    parent: tk.Widget, text: str, command, bg: str, size: int = 10
) -> tk.Button:
    h = bg.lstrip("#")
    r, g, b = (max(0, int(h[i: i + 2], 16) - 28) for i in (0, 2, 4))
    darker = f"#{r:02X}{g:02X}{b:02X}"
    return tk.Button(
        parent,
        text=text,
        command=command,
        font=("Segoe UI", size, "bold"),
        bg=bg, fg="white",
        relief="flat", bd=0,
        padx=14, pady=7,
        cursor="hand2",
        activebackground=darker,
        activeforeground="white",
    )


def _labelframe(
    parent: tk.Widget, title: str, row: int, expand: bool = False,
    accent: str = C_ACCENT
) -> tk.LabelFrame:
    lf = tk.LabelFrame(
        parent, text=title,
        font=("Segoe UI", 9, "bold"),
        bg=C_CARD, fg=accent,
        relief="solid", bd=1,
        padx=10, pady=8,
    )
    sticky = "nsew" if expand else "ew"
    lf.grid(row=row, column=0, sticky=sticky, pady=(0, 6))
    if expand:
        parent.rowconfigure(row, weight=1)
    return lf


# ══════════════════════════════════════════════════════════════
#  ABA 2 — PLANILHA DE CONTROLE
# ══════════════════════════════════════════════════════════════
class PlanilhaTab(tk.Frame):
    """Aba de integração com a planilha Excel de controle de certidões."""

    # Mapeamento de colunas: nome amigável → (keywords exatas, keywords parciais)
    # Ordem de prioridade: match exato do cabeçalho > match parcial
    # Colunas reais da planilha (linha 2 é o cabeçalho):
    #   B=SOLIC. DIA, C=FORNECEDOR, D=?, E=CPF/CNPJ, F=TIPO, G=MUNICÍPIO,
    #   H=SOLICITANTE, I=PROTOCOLO (P), J=VALOR PAGO (R$), K=?, L=PAGO DIA,
    #   M=PRAZOS DO CRI, N=STATUS
    _COL_MAP = {
        "SOLIC. DIA":   (["solic. dia", "solic.dia"],         ["solic"]),
        "FORNECEDOR":   (["fornecedor"],                      ["fornec"]),
        "CPF/CNPJ":     (["cpf/cnpj", "cpf / cnpj"],         ["cpf", "cnpj"]),
        "TIPO":         (["tipo"],                            []),
        "MUNICÍPIO":    (["município", "municipio"],          ["munic"]),
        "SOLICITANTE":  (["solicitante"],                     []),
        "PROTOCOLO":    (["protocolo (p)", "protocolo"],     ["prot"]),
        "VALOR PAGO":   (["valor pago (r$)", "valor pago"],   ["valor"]),
        "PAGO DIA":     (["pago dia", "prazos do cri"],       ["pago"]),
        "STATUS":       (["status"],                          ["situac"]),
    }

    def __init__(self, parent: tk.Widget, status_var: tk.StringVar) -> None:
        super().__init__(parent, bg=C_BG)
        self._status_var = status_var
        self._rows: list[dict] = []          # dados carregados
        self._filtered: list[dict] = []      # dados filtrados
        self._col_names: list[str] = []      # colunas detectadas
        self._planilha_path: str = ""

        self._build_ui()
        # Tenta carregar automaticamente ao iniciar
        self.after(300, self._auto_load)

    # ──────────────────────────────────────────────────────────
    #  CONSTRUÇÃO DO LAYOUT
    # ──────────────────────────────────────────────────────────
    def _build_ui(self) -> None:
        self.columnconfigure(0, weight=1)
        self.rowconfigure(3, weight=1)

        self._build_load_bar()      # row 0
        self._build_filter_bar()    # row 1
        self._build_counter_bar()   # row 2
        self._build_results_area()  # row 3
        self._build_detail_card()   # row 4

    # ── Barra de carregamento ─────────────────────────────────
    def _build_load_bar(self) -> None:
        card = tk.LabelFrame(
            self, text=" 📊  Planilha de Controle ",
            font=("Segoe UI", 9, "bold"),
            bg=C_CARD, fg=C_EXCEL,
            relief="solid", bd=1,
            padx=10, pady=8,
        )
        card.grid(row=0, column=0, sticky="ew", pady=(0, 6))

        inner = tk.Frame(card, bg=C_CARD)
        inner.pack(fill=tk.X)
        inner.columnconfigure(0, weight=1)

        self._path_var = tk.StringVar(value="(nenhuma planilha carregada)")
        tk.Entry(
            inner,
            textvariable=self._path_var,
            font=("Segoe UI", 9),
            state="readonly",
            readonlybackground="#FAFAFA",
            relief="solid", bd=1,
            cursor="arrow",
        ).grid(row=0, column=0, sticky="ew", ipady=5, padx=(0, 8))

        _make_btn(inner, "📂  Selecionar", self._on_select_planilha, C_EXCEL, size=9
                  ).grid(row=0, column=1, padx=(0, 6))
        _make_btn(inner, "🔄  Recarregar", self._on_reload, C_EXCEL_DARK, size=9
                  ).grid(row=0, column=2)

        tk.Label(
            card,
            text="Detecta automaticamente a planilha no OneDrive. Use 'Selecionar' para escolher outro arquivo.",
            font=("Segoe UI", 8),
            bg=C_CARD, fg=C_MUTED,
        ).pack(anchor=tk.W, pady=(6, 0))

    # ── Barra de filtros ──────────────────────────────────────
    def _build_filter_bar(self) -> None:
        card = tk.LabelFrame(
            self, text=" 🔍  Filtros ",
            font=("Segoe UI", 9, "bold"),
            bg=C_CARD, fg=C_EXCEL,
            relief="solid", bd=1,
            padx=10, pady=8,
        )
        card.grid(row=1, column=0, sticky="ew", pady=(0, 6))

        inner = tk.Frame(card, bg=C_CARD)
        inner.pack(fill=tk.X)

        # Campo de busca
        tk.Label(inner, text="Buscar:", font=("Segoe UI", 9, "bold"),
                 bg=C_CARD, fg="#333").pack(side=tk.LEFT, padx=(0, 6))

        self._search_var = tk.StringVar()
        self._search_var.trace_add("write", lambda *_: self._apply_filters())
        tk.Entry(
            inner,
            textvariable=self._search_var,
            font=("Segoe UI", 10),
            relief="solid", bd=1,
            width=30,
        ).pack(side=tk.LEFT, ipady=5, padx=(0, 16))

        tk.Label(inner, text="(Protocolo, CPF/CNPJ ou Fornecedor)",
                 font=("Segoe UI", 8), bg=C_CARD, fg=C_MUTED).pack(side=tk.LEFT, padx=(0, 20))

        # Filtro por tipo
        tk.Label(inner, text="Tipo:", font=("Segoe UI", 9, "bold"),
                 bg=C_CARD, fg="#333").pack(side=tk.LEFT, padx=(0, 6))

        self._tipo_var = tk.StringVar(value="Todos")
        self._tipo_combo = ttk.Combobox(
            inner,
            textvariable=self._tipo_var,
            values=["Todos", "PENHOR", "FIDUCIARIA", "INT.TEOR"],
            state="readonly",
            width=14,
            font=("Segoe UI", 9),
        )
        self._tipo_combo.pack(side=tk.LEFT, ipady=3, padx=(0, 10))
        self._tipo_combo.bind("<<ComboboxSelected>>", lambda *_: self._apply_filters())

        _make_btn(inner, "✖  Limpar Filtros", self._on_clear_filters, C_MUTED, size=9
                  ).pack(side=tk.LEFT, padx=(10, 0))

    # ── Barra de contadores ───────────────────────────────────
    def _build_counter_bar(self) -> None:
        bar = tk.Frame(self, bg=C_COUNTER_BG, relief="solid", bd=1)
        bar.grid(row=2, column=0, sticky="ew", pady=(0, 6))

        inner = tk.Frame(bar, bg=C_COUNTER_BG, pady=6)
        inner.pack()

        lbl_kw = dict(font=("Segoe UI", 10, "bold"), bg=C_COUNTER_BG)
        sep_kw = dict(text="│", bg=C_COUNTER_BG, fg="#B0BEC5", font=("Segoe UI", 14))

        self._lbl_total_pl = tk.Label(inner, text="Total na planilha: 0",
                                      fg="#1A3C5E", **lbl_kw)
        self._lbl_total_pl.pack(side=tk.LEFT, padx=14)

        tk.Label(inner, **sep_kw).pack(side=tk.LEFT)

        self._lbl_result_pl = tk.Label(inner, text="Resultados: 0",
                                       fg=C_EXCEL, **lbl_kw)
        self._lbl_result_pl.pack(side=tk.LEFT, padx=14)

        tk.Label(inner, **sep_kw).pack(side=tk.LEFT)

        self._lbl_status_pl = tk.Label(inner, text="Planilha não carregada",
                                       fg=C_MUTED, **lbl_kw)
        self._lbl_status_pl.pack(side=tk.LEFT, padx=14)

    # ── Área de resultados (Treeview) ─────────────────────────
    def _build_results_area(self) -> None:
        card = tk.LabelFrame(
            self, text=" 📋  Certidões ",
            font=("Segoe UI", 9, "bold"),
            bg=C_CARD, fg=C_EXCEL,
            relief="solid", bd=1,
            padx=10, pady=8,
        )
        card.grid(row=3, column=0, sticky="nsew", pady=(0, 6))
        self.rowconfigure(3, weight=1)

        frm = tk.Frame(card, bg=C_CARD)
        frm.pack(fill=tk.BOTH, expand=True)

        # Colunas visíveis na tabela
        self._tree_cols = (
            "PROTOCOLO", "FORNECEDOR", "CPF/CNPJ", "TIPO",
            "MUNICÍPIO", "SOLIC. DI", "VALOR PAGO", "STATUS"
        )
        self._pl_tree = ttk.Treeview(
            frm, columns=self._tree_cols, show="headings", selectmode="browse"
        )

        col_widths = {
            "PROTOCOLO": 140, "FORNECEDOR": 200, "CPF/CNPJ": 130,
            "TIPO": 100, "MUNICÍPIO": 120, "SOLIC. DI": 90,
            "VALOR PAGO": 100, "STATUS": 110,
        }
        for col in self._tree_cols:
            self._pl_tree.heading(col, text=col, anchor=tk.W)
            self._pl_tree.column(col, width=col_widths.get(col, 120),
                                 minwidth=80, stretch=(col == "FORNECEDOR"))

        # Tags de cor por status
        self._pl_tree.tag_configure("ok",       foreground=C_SUCCESS)
        self._pl_tree.tag_configure("pendente",  foreground="#E65100")
        self._pl_tree.tag_configure("cancelado", foreground=C_DANGER)
        self._pl_tree.tag_configure("normal",    foreground="#212121")
        self._pl_tree.tag_configure("odd",       background="#F7F9FC")

        sy = ttk.Scrollbar(frm, orient=tk.VERTICAL,   command=self._pl_tree.yview)
        sx = ttk.Scrollbar(frm, orient=tk.HORIZONTAL, command=self._pl_tree.xview)
        self._pl_tree.configure(yscrollcommand=sy.set, xscrollcommand=sx.set)

        sy.pack(side=tk.RIGHT,  fill=tk.Y)
        sx.pack(side=tk.BOTTOM, fill=tk.X)
        self._pl_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self._pl_tree.bind("<<TreeviewSelect>>", self._on_pl_select)

    # ── Card de detalhes ──────────────────────────────────────
    def _build_detail_card(self) -> None:
        self._detail_frame = tk.LabelFrame(
            self, text=" 📄  Detalhes da Certidão Selecionada ",
            font=("Segoe UI", 9, "bold"),
            bg=C_CARD, fg=C_EXCEL,
            relief="solid", bd=1,
            padx=12, pady=10,
        )
        self._detail_frame.grid(row=4, column=0, sticky="ew", pady=(0, 6))

        # Grid de campos de detalhe
        self._detail_vars: dict[str, tk.StringVar] = {}
        detail_fields = [
            ("PROTOCOLO",  "FORNECEDOR"),
            ("CPF/CNPJ",   "TIPO"),
            ("MUNICÍPIO",  "SOLIC. DI"),
            ("SOLICITANTE","VALOR PAGO"),
            ("PAGO DIA",   "STATUS"),
        ]

        for row_idx, (left, right) in enumerate(detail_fields):
            for col_idx, field in enumerate([left, right]):
                lbl_col = col_idx * 4
                val_col = col_idx * 4 + 1

                tk.Label(
                    self._detail_frame,
                    text=f"{field}:",
                    font=("Segoe UI", 9, "bold"),
                    bg=C_CARD, fg=C_EXCEL,
                    anchor=tk.E,
                ).grid(row=row_idx, column=lbl_col, sticky="e", padx=(8, 4), pady=3)

                var = tk.StringVar(value="—")
                self._detail_vars[field] = var
                tk.Label(
                    self._detail_frame,
                    textvariable=var,
                    font=("Segoe UI", 9),
                    bg=C_CARD, fg="#212121",
                    anchor=tk.W,
                    width=28,
                ).grid(row=row_idx, column=val_col, sticky="w", padx=(0, 16), pady=3)

            # Separador entre colunas
            if col_idx == 0:
                tk.Frame(self._detail_frame, bg="#E0E0E0", width=1
                         ).grid(row=row_idx, column=2, sticky="ns", padx=8)

        self._detail_frame.columnconfigure(1, weight=1)
        self._detail_frame.columnconfigure(5, weight=1)

        # Mensagem inicial
        self._detail_placeholder = tk.Label(
            self._detail_frame,
            text="Selecione uma linha na tabela para ver os detalhes da certidão.",
            font=("Segoe UI", 9, "italic"),
            bg=C_CARD, fg=C_MUTED,
        )
        self._detail_placeholder.grid(row=0, column=0, columnspan=8, pady=8)
        self._hide_detail_fields()

    def _hide_detail_fields(self) -> None:
        for widget in self._detail_frame.winfo_children():
            if widget != self._detail_placeholder:
                widget.grid_remove()
        self._detail_placeholder.grid()

    def _show_detail_fields(self) -> None:
        self._detail_placeholder.grid_remove()
        for widget in self._detail_frame.winfo_children():
            if widget != self._detail_placeholder:
                widget.grid()

    # ──────────────────────────────────────────────────────────
    #  CARREGAMENTO DA PLANILHA
    # ──────────────────────────────────────────────────────────
    def _auto_load(self) -> None:
        path = _find_planilha()
        if path:
            self._load_planilha(path)
        else:
            self._lbl_status_pl.config(
                text="Planilha não encontrada automaticamente", fg=C_DANGER
            )
            self._status_var.set(
                "Planilha de controle não localizada. Use 'Selecionar' para escolher o arquivo."
            )

    def _on_select_planilha(self) -> None:
        path = filedialog.askopenfilename(
            title="Selecionar Planilha de Controle",
            filetypes=[("Excel", "*.xlsx *.xlsm *.xls"), ("Todos", "*.*")],
        )
        if path:
            self._load_planilha(path)

    def _on_reload(self) -> None:
        if self._planilha_path and os.path.isfile(self._planilha_path):
            self._load_planilha(self._planilha_path)
        else:
            self._auto_load()

    def _load_planilha(self, path: str) -> None:
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror(
                "Dependência ausente",
                "A biblioteca 'openpyxl' não está instalada.\n\n"
                "Instale com:\n    pip install openpyxl\n\n"
                "Após instalar, reinicie o programa.",
            )
            return

        self._planilha_path = path
        self._path_var.set(path)
        self._lbl_status_pl.config(text="Carregando...", fg=C_MUTED)
        self._status_var.set(f"Carregando planilha: {os.path.basename(path)}")
        self.update_idletasks()

        threading.Thread(
            target=self._load_worker,
            args=(path,),
            daemon=True,
        ).start()

    def _load_worker(self, path: str) -> None:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

            # Tenta encontrar aba 'CTRL' especificamente, ou usa a primeira aba
            ws = None
            for sheet_name in wb.sheetnames:
                if sheet_name.upper() == "CTRL":
                    ws = wb[sheet_name]
                    break
            # Fallback: usa a primeira aba
            if ws is None:
                ws = wb.worksheets[0]

            rows_iter = ws.iter_rows(values_only=True)

            # Pula a linha 1 (título) e lê cabeçalho na linha 2
            next(rows_iter, None)  # skip linha 1
            header_row = next(rows_iter, None)
            if header_row is None:
                self.after(0, self._load_error, "Planilha vazia.")
                return

            col_index = self._detect_columns(header_row)

            rows = []
            for raw_row in rows_iter:
                if all(v is None for v in raw_row):
                    continue
                record = {}
                for col_name, idx in col_index.items():
                    if idx < len(raw_row):
                        val = raw_row[idx]
                        record[col_name] = str(val).strip() if val is not None else ""
                    else:
                        record[col_name] = ""
                rows.append(record)

            wb.close()
            self.after(0, self._load_done, rows, col_index)

        except Exception as exc:
            self.after(0, self._load_error, str(exc))

    def _detect_columns(self, header_row: tuple) -> dict[str, int]:
        """Detecta índices das colunas pelo cabeçalho.

        Estratégia em dois passos:
        1. Match exato: verifica se alguma keyword exata é igual ao valor da célula.
        2. Match parcial: verifica se alguma keyword parcial está contida no valor.
        Isso evita que 'pago' (parcial) bata em 'VALOR PAGO' antes de 'PAGO DIA'.
        """
        col_index: dict[str, int] = {}

        # Pré-processa o cabeçalho uma vez
        header_lower = [
            str(v).lower().strip() if v is not None else ""
            for v in header_row
        ]

        for col_name, (exact_kws, partial_kws) in self._COL_MAP.items():
            found_idx = None

            # Passo 1: match exato (célula == keyword)
            for idx, cell_str in enumerate(header_lower):
                if cell_str in exact_kws:
                    found_idx = idx
                    break

            # Passo 2: match parcial (keyword contida na célula)
            if found_idx is None and partial_kws:
                for idx, cell_str in enumerate(header_lower):
                    if cell_str and any(kw in cell_str for kw in partial_kws):
                        found_idx = idx
                        break

            if found_idx is not None:
                col_index[col_name] = found_idx

        return col_index

    def _load_done(self, rows: list, col_index: dict) -> None:
        self._rows = rows
        self._col_names = list(col_index.keys())
        self._filtered = list(rows)

        total = len(rows)
        self._lbl_total_pl.config(text=f"Total na planilha: {total}")
        self._lbl_status_pl.config(
            text=f"Carregada: {os.path.basename(self._planilha_path)}", fg=C_EXCEL
        )
        self._status_var.set(
            f"Planilha carregada: {total} registro(s) — {os.path.basename(self._planilha_path)}"
        )
        self._render_tree(rows)

    def _load_error(self, msg: str) -> None:
        self._lbl_status_pl.config(text="Erro ao carregar", fg=C_DANGER)
        self._status_var.set(f"Erro ao carregar planilha: {msg}")
        messagebox.showerror("Erro ao carregar planilha", msg)

    # ──────────────────────────────────────────────────────────
    #  FILTROS
    # ──────────────────────────────────────────────────────────
    def _apply_filters(self) -> None:
        search = self._search_var.get().strip().lower()
        tipo   = self._tipo_var.get()

        filtered = []
        for row in self._rows:
            # Filtro por tipo
            if tipo != "Todos":
                row_tipo = row.get("TIPO", "").upper()
                if tipo not in row_tipo:
                    continue

            # Filtro por texto (protocolo, CPF/CNPJ, fornecedor)
            if search:
                proto     = row.get("PROTOCOLO", "").lower()
                cpf       = row.get("CPF/CNPJ", "").lower()
                fornec    = row.get("FORNECEDOR", "").lower()
                if search not in proto and search not in cpf and search not in fornec:
                    continue

            filtered.append(row)

        self._filtered = filtered
        self._render_tree(filtered)

    def _on_clear_filters(self) -> None:
        self._search_var.set("")
        self._tipo_var.set("Todos")
        self._apply_filters()

    # ──────────────────────────────────────────────────────────
    #  RENDERIZAÇÃO DA TABELA
    # ──────────────────────────────────────────────────────────
    def _render_tree(self, rows: list) -> None:
        self._pl_tree.delete(*self._pl_tree.get_children())

        for i, row in enumerate(rows):
            values = tuple(row.get(col, "") for col in self._tree_cols)
            tag = self._status_tag(row.get("STATUS", ""))
            if i % 2 == 1:
                tag = tag + "_odd" if tag != "normal" else "odd"
            self._pl_tree.insert("", tk.END, values=values, tags=(tag,))

        self._lbl_result_pl.config(text=f"Resultados: {len(rows)}")
        self._hide_detail_fields()

    @staticmethod
    def _status_tag(status: str) -> str:
        s = status.upper()
        if any(k in s for k in ["OK", "EMITID", "CONCLU", "PAGO"]):
            return "ok"
        if any(k in s for k in ["CANCEL", "RECUS", "NEGAD"]):
            return "cancelado"
        if any(k in s for k in ["PEND", "AGUARD", "ANDAMENTO"]):
            return "pendente"
        return "normal"

    # ──────────────────────────────────────────────────────────
    #  SELEÇÃO NA TABELA → CARD DE DETALHES
    # ──────────────────────────────────────────────────────────
    def _on_pl_select(self, _event=None) -> None:
        sel = self._pl_tree.selection()
        if not sel:
            self._hide_detail_fields()
            return

        iid = sel[0]
        idx = self._pl_tree.index(iid)
        if idx >= len(self._filtered):
            return

        row = self._filtered[idx]
        self._show_detail_fields()

        all_fields = [
            "PROTOCOLO", "FORNECEDOR", "CPF/CNPJ", "TIPO",
            "MUNICÍPIO", "SOLIC. DI", "SOLICITANTE",
            "VALOR PAGO", "PAGO DIA", "STATUS",
        ]
        for field in all_fields:
            if field in self._detail_vars:
                val = row.get(field, "—") or "—"
                self._detail_vars[field].set(val)


# ══════════════════════════════════════════════════════════════
#  CLASSE PRINCIPAL (STANDALONE) — mantida para uso independente
# ══════════════════════════════════════════════════════════════
class ConsultaCertidoesApp(tk.Tk):
    """Janela principal standalone – Consulta Certidões ONR (sem abas)."""

    _PLACEHOLDER = "P26070520844D\nP26070520845D\nP26070520846D"

    def __init__(self) -> None:
        super().__init__()

        try:
            from ctypes import windll
            windll.shcore.SetProcessDpiAwareness(1)
        except Exception:
            pass

        self.title("Consulta Certidões ONR")
        self.geometry("1040x780")
        self.minsize(820, 600)
        self.configure(bg=C_BG)

        try:
            self.iconbitmap("icon.ico")
        except Exception:
            pass

        self._found_files: dict = {}
        self._drive_files: set  = set()
        self._ph_active: bool = True
        self._temp_dirs: list = []

        self.protocol("WM_DELETE_WINDOW", self._on_close)

        self._apply_styles()
        self._build_ui()
        self._center_window()

    def _apply_styles(self) -> None:
        s = ttk.Style(self)
        s.theme_use("clam")
        s.configure("Treeview", font=("Segoe UI", 10), rowheight=30,
                    background=C_CARD, fieldbackground=C_CARD, foreground="#212121")
        s.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"),
                    background=C_ACCENT, foreground="white", relief="flat")
        s.map("Treeview.Heading", background=[("active", "#0D47A1")])
        s.map("Treeview",
              background=[("selected", "#BBDEFB")],
              foreground=[("selected", "#0D47A1")])

    def _build_ui(self) -> None:
        self._build_header()
        body = tk.Frame(self, bg=C_BG, padx=18, pady=12)
        body.pack(fill=tk.BOTH, expand=True)
        body.columnconfigure(0, weight=1)
        body.rowconfigure(5, weight=1)
        self._build_folder_row(body)
        self._build_drive_row(body)
        self._build_protocol_box(body)
        self._build_action_bar(body)
        self._build_counter_bar(body)
        self._build_results_panel(body)
        self._build_status_bar()

    def _build_header(self) -> None:
        hdr = tk.Frame(self, bg=C_HEADER)
        hdr.pack(fill=tk.X)
        tk.Label(hdr, text="🔍  Consulta Certidões ONR",
                 font=("Segoe UI", 17, "bold"), bg=C_HEADER, fg="white", pady=12).pack()
        tk.Label(hdr, text="Busca local + OneDrive — baixe certidões do drive para sua pasta",
                 font=("Segoe UI", 9), bg=C_HEADER, fg="#90CAF9").pack(pady=(0, 2))
        tk.Frame(hdr, bg="#0D47A1", height=3).pack(fill=tk.X, pady=(6, 0))

    def _build_folder_row(self, body: tk.Frame) -> None:
        card = _labelframe(body, " 📁  Pasta Local de Busca ", row=0)
        inner = tk.Frame(card, bg=C_CARD)
        inner.pack(fill=tk.X)
        inner.columnconfigure(0, weight=1)
        self._folder_var = tk.StringVar()
        tk.Entry(inner, textvariable=self._folder_var, font=("Segoe UI", 10),
                 state="readonly", readonlybackground="#FAFAFA",
                 relief="solid", bd=1, cursor="arrow"
                 ).grid(row=0, column=0, sticky="ew", ipady=6, padx=(0, 8))
        _make_btn(inner, "📂  Selecionar Pasta", self._on_select_folder, C_ACCENT
                  ).grid(row=0, column=1)

    def _build_drive_row(self, body: tk.Frame) -> None:
        card = _labelframe(body, " ☁  Pasta OneDrive (fallback) ", row=1)
        inner = tk.Frame(card, bg=C_CARD)
        inner.pack(fill=tk.X)
        inner.columnconfigure(0, weight=1)
        self._drive_var = tk.StringVar(value=ONEDRIVE_DEFAULT)
        tk.Entry(inner, textvariable=self._drive_var, font=("Segoe UI", 9),
                 relief="solid", bd=1
                 ).grid(row=0, column=0, sticky="ew", ipady=5, padx=(0, 8))
        _make_btn(inner, "📂  Selecionar", self._on_select_drive, C_DRIVE, size=9
                  ).grid(row=0, column=1, padx=(0, 6))
        _make_btn(inner, "🌐  Abrir no Browser", self._on_open_drive_web, C_MUTED, size=9
                  ).grid(row=0, column=2)
        self._drive_enabled = tk.BooleanVar(value=True)
        tk.Checkbutton(card, text="Buscar no OneDrive quando não encontrado localmente",
                       variable=self._drive_enabled, font=("Segoe UI", 9),
                       bg=C_CARD, fg=C_MUTED, activebackground=C_CARD
                       ).pack(anchor=tk.W, pady=(6, 0))

    def _build_protocol_box(self, body: tk.Frame) -> None:
        card = _labelframe(body, " 📋  Protocolos para Consultar ", row=2)
        tk.Label(card, text="Cole os protocolos abaixo, um por linha  (a busca ignora maiúsculas/minúsculas):",
                 font=("Segoe UI", 9), bg=C_CARD, fg=C_MUTED).pack(anchor=tk.W, pady=(0, 5))
        frm = tk.Frame(card, bg=C_CARD)
        frm.pack(fill=tk.BOTH, expand=True)
        self._proto_txt = tk.Text(frm, height=5, font=("Courier New", 11),
                                  relief="solid", bd=1, wrap=tk.NONE, fg="gray", undo=True)
        sy = ttk.Scrollbar(frm, orient=tk.VERTICAL,   command=self._proto_txt.yview)
        sx = ttk.Scrollbar(frm, orient=tk.HORIZONTAL, command=self._proto_txt.xview)
        self._proto_txt.configure(yscrollcommand=sy.set, xscrollcommand=sx.set)
        sy.pack(side=tk.RIGHT,  fill=tk.Y)
        sx.pack(side=tk.BOTTOM, fill=tk.X)
        self._proto_txt.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self._proto_txt.insert("1.0", self._PLACEHOLDER)
        self._proto_txt.bind("<FocusIn>",  self._ph_clear)
        self._proto_txt.bind("<FocusOut>", self._ph_restore)

    def _build_action_bar(self, body: tk.Frame) -> None:
        bar = tk.Frame(body, bg=C_BG, pady=7)
        bar.grid(row=3, column=0, sticky="ew")
        self._btn_consult = _make_btn(bar, "🔍   CONSULTAR", self._on_consult, C_ACCENT, size=11)
        self._btn_consult.pack(side=tk.LEFT, padx=(0, 8))
        _make_btn(bar, "🗑   LIMPAR", self._on_clear, C_MUTED, size=11).pack(side=tk.LEFT, padx=(0, 8))
        self._btn_open = _make_btn(bar, "📂   ABRIR", self._on_open_file, C_SUCCESS, size=11)
        self._btn_open.pack(side=tk.LEFT, padx=(0, 8))
        self._btn_open.config(state="disabled")
        self._btn_print = _make_btn(bar, "🖨   IMPRIMIR", self._on_print_file, C_PRINT, size=11)
        self._btn_print.pack(side=tk.LEFT, padx=(0, 8))
        self._btn_print.config(state="disabled")
        self._btn_download = _make_btn(bar, "⬇   BAIXAR DO DRIVE", self._on_download_from_drive, C_DRIVE, size=11)
        self._btn_download.pack(side=tk.LEFT)
        self._btn_download.config(state="disabled")

    def _build_counter_bar(self, body: tk.Frame) -> None:
        bar = tk.Frame(body, bg=C_COUNTER_BG, relief="solid", bd=1)
        bar.grid(row=4, column=0, sticky="ew", pady=(0, 6))
        inner = tk.Frame(bar, bg=C_COUNTER_BG, pady=7)
        inner.pack()
        lbl_kw = dict(font=("Segoe UI", 10, "bold"), bg=C_COUNTER_BG)
        sep_kw = dict(text="│", bg=C_COUNTER_BG, fg="#B0BEC5", font=("Segoe UI", 14))
        self._lbl_total    = tk.Label(inner, text="Total: 0",              fg="#1A3C5E", **lbl_kw)
        self._lbl_total.pack(side=tk.LEFT, padx=14)
        tk.Label(inner, **sep_kw).pack(side=tk.LEFT)
        self._lbl_found    = tk.Label(inner, text="✅  Local: 0",          fg=C_SUCCESS, **lbl_kw)
        self._lbl_found.pack(side=tk.LEFT, padx=14)
        tk.Label(inner, **sep_kw).pack(side=tk.LEFT)
        self._lbl_drive    = tk.Label(inner, text="☁  Drive: 0",           fg=C_DRIVE,   **lbl_kw)
        self._lbl_drive.pack(side=tk.LEFT, padx=14)
        tk.Label(inner, **sep_kw).pack(side=tk.LEFT)
        self._lbl_notfound = tk.Label(inner, text="❌  Não encontrados: 0", fg=C_DANGER,  **lbl_kw)
        self._lbl_notfound.pack(side=tk.LEFT, padx=14)

    def _build_results_panel(self, body: tk.Frame) -> None:
        card = _labelframe(body, " 📊  Resultados ", row=5, expand=True)
        frm = tk.Frame(card, bg=C_CARD)
        frm.pack(fill=tk.BOTH, expand=True)
        cols = ("status", "protocolo", "arquivo", "caminho")
        self._tree = ttk.Treeview(frm, columns=cols, show="headings", selectmode="extended")
        self._tree.heading("status",    text="Status",             anchor=tk.CENTER)
        self._tree.heading("protocolo", text="Protocolo",          anchor=tk.W)
        self._tree.heading("arquivo",   text="Arquivo encontrado", anchor=tk.W)
        self._tree.heading("caminho",   text="Caminho completo",   anchor=tk.W)
        self._tree.column("status",    width=185, minwidth=150, anchor=tk.CENTER, stretch=False)
        self._tree.column("protocolo", width=160, minwidth=130, stretch=False)
        self._tree.column("arquivo",   width=280, minwidth=180)
        self._tree.column("caminho",   width=360, minwidth=200)
        self._tree.tag_configure("found",    foreground=C_SUCCESS)
        self._tree.tag_configure("drive",    foreground=C_DRIVE)
        self._tree.tag_configure("notfound", foreground=C_DANGER)
        sy = ttk.Scrollbar(frm, orient=tk.VERTICAL,   command=self._tree.yview)
        sx = ttk.Scrollbar(frm, orient=tk.HORIZONTAL, command=self._tree.xview)
        self._tree.configure(yscrollcommand=sy.set, xscrollcommand=sx.set)
        sy.pack(side=tk.RIGHT,  fill=tk.Y)
        sx.pack(side=tk.BOTTOM, fill=tk.X)
        self._tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self._tree.bind("<<TreeviewSelect>>", self._on_tree_select)
        self._tree.bind("<Double-1>",         lambda _e: self._on_open_file())

    def _build_status_bar(self) -> None:
        self._status_var = tk.StringVar(value="Pronto para consultar.")
        tk.Label(self, textvariable=self._status_var, font=("Segoe UI", 9),
                 bg=C_STATUS_BG, fg="#455A64", anchor=tk.W, padx=10, pady=4,
                 relief="sunken").pack(side=tk.BOTTOM, fill=tk.X)

    # ── Placeholder ───────────────────────────────────────────
    def _ph_clear(self, _event=None) -> None:
        if self._ph_active:
            self._proto_txt.delete("1.0", tk.END)
            self._proto_txt.config(fg="#212121")
            self._ph_active = False

    def _ph_restore(self, _event=None) -> None:
        if not self._proto_txt.get("1.0", tk.END).strip():
            self._proto_txt.insert("1.0", self._PLACEHOLDER)
            self._proto_txt.config(fg="gray")
            self._ph_active = True

    # ── Ações ─────────────────────────────────────────────────
    def _on_select_folder(self) -> None:
        path = filedialog.askdirectory(title="Selecionar pasta local de certidões")
        if path:
            self._folder_var.set(path)
            self._status_var.set(f"Pasta local selecionada: {path}")

    def _on_select_drive(self) -> None:
        path = filedialog.askdirectory(title="Selecionar pasta do OneDrive")
        if path:
            self._drive_var.set(path)
            self._status_var.set(f"Pasta OneDrive selecionada: {path}")

    def _on_open_drive_web(self) -> None:
        try:
            import webbrowser
            webbrowser.open(ONEDRIVE_WEB_URL)
        except Exception as exc:
            messagebox.showerror("Erro", str(exc))

    def _on_consult(self) -> None:
        folder     = self._folder_var.get().strip()
        drive_path = self._drive_var.get().strip()
        use_drive  = self._drive_enabled.get()
        protocols  = self._get_protocols()

        if not folder:
            messagebox.showwarning("Atenção", "Selecione a pasta local antes de consultar.")
            return
        if not os.path.isdir(folder):
            messagebox.showerror("Erro", f"Pasta local não encontrada:\n{folder}")
            return
        if not protocols:
            messagebox.showwarning("Atenção", "Informe ao menos um protocolo.")
            return

        if use_drive and drive_path and not os.path.isdir(drive_path):
            resp = messagebox.askyesno(
                "Pasta OneDrive não encontrada",
                f"A pasta OneDrive não foi localizada:\n{drive_path}\n\n"
                "Continuar apenas com a busca local?",
            )
            if not resp:
                return
            use_drive = False

        self._tree.delete(*self._tree.get_children())
        self._found_files.clear()
        self._drive_files.clear()
        self._btn_open.config(state="disabled")
        self._btn_print.config(state="disabled")
        self._btn_download.config(state="disabled")
        self._update_counters(0, 0, 0, 0)

        self._btn_consult.config(state="disabled", text="⏳  Consultando...")
        self._status_var.set("Indexando arquivos … aguarde")
        self.update_idletasks()

        threading.Thread(
            target=self._search_worker,
            args=(folder, drive_path if use_drive else "", protocols),
            daemon=True,
        ).start()

    def _search_worker(self, local_folder: str, drive_folder: str, protocols: list) -> None:
        def index_folder(folder: str) -> list[tuple[str, str]]:
            files = []
            for dirpath, _dirs, fnames in os.walk(folder):
                for fname in fnames:
                    files.append((fname, os.path.join(dirpath, fname)))
            return files

        local_files = index_folder(local_folder)
        drive_files = index_folder(drive_folder) if drive_folder else []

        results = []
        local_count = drive_count = not_found_count = 0

        for proto in protocols:
            proto_lower = proto.lower()
            local_matches = [(f, p) for f, p in local_files if proto_lower in f.lower()]
            if local_matches:
                local_count += 1
                for fname, fpath in local_matches:
                    results.append(("found", proto, fname, fpath))
                continue
            drive_matches = [(f, p) for f, p in drive_files if proto_lower in f.lower()]
            if drive_matches:
                drive_count += 1
                for fname, fpath in drive_matches:
                    results.append(("drive", proto, fname, fpath))
            else:
                not_found_count += 1
                results.append(("notfound", proto, "", ""))

        self.after(0, self._show_results, results, len(protocols),
                   local_count, drive_count, not_found_count)

    def _show_results(self, results, total, local, drive, not_found) -> None:
        for status, proto, fname, fpath in results:
            if status == "found":
                iid = self._tree.insert("", tk.END,
                                        values=("✅  Local", proto, fname, fpath),
                                        tags=("found",))
                self._found_files[iid] = fpath
            elif status == "drive":
                iid = self._tree.insert("", tk.END,
                                        values=("☁  OneDrive", proto, fname, fpath),
                                        tags=("drive",))
                self._found_files[iid] = fpath
                self._drive_files.add(iid)
            else:
                self._tree.insert("", tk.END,
                                  values=("❌  Não encontrado", proto, "—", "—"),
                                  tags=("notfound",))

        self._update_counters(total, local, drive, not_found)
        self._btn_consult.config(state="normal", text="🔍   CONSULTAR")
        self._status_var.set(
            f"Consulta finalizada  •  {total} protocolo(s)  •  "
            f"{local} local / {drive} OneDrive / {not_found} não encontrado(s)"
        )

    def _on_tree_select(self, _event=None) -> None:
        sel = self._tree.selection()
        found_sel = [iid for iid in sel if iid in self._found_files]
        drive_sel = [iid for iid in found_sel if iid in self._drive_files]
        self._btn_open.config(state="normal" if found_sel else "disabled")
        if found_sel:
            n = len(found_sel)
            self._btn_print.config(state="normal",
                                   text=f"🖨  IMPRIMIR ({n})" if n > 1 else "🖨   IMPRIMIR")
        else:
            self._btn_print.config(state="disabled", text="🖨   IMPRIMIR")
        if drive_sel:
            n = len(drive_sel)
            self._btn_download.config(state="normal",
                                      text=f"⬇   BAIXAR DO DRIVE ({n})" if n > 1 else "⬇   BAIXAR DO DRIVE")
        else:
            self._btn_download.config(state="disabled", text="⬇   BAIXAR DO DRIVE")

    def _on_open_file(self) -> None:
        sel = self._tree.selection()
        if not sel:
            return
        iid = sel[0]
        if iid not in self._found_files:
            messagebox.showinfo("Aviso", "Selecione um item ✅/☁ para abrir.")
            return
        path = self._found_files[iid]
        if not os.path.exists(path):
            messagebox.showerror("Arquivo não encontrado", f"Não localizado:\n{path}")
            return
        try:
            os.startfile(path)
        except Exception as exc:
            messagebox.showerror("Erro ao abrir arquivo", str(exc))

    def _open_from_zip(self, zip_path: str) -> None:
        """Extrai PDFs de um ZIP e abre o PDF extraído. Se múltiplos, mostra diálogo."""
        try:
            with zipfile.ZipFile(zip_path, "r") as zf:
                pdf_names = [n for n in zf.namelist()
                             if n.lower().endswith(".pdf") and not n.startswith("__MACOSX")]
        except zipfile.BadZipFile:
            messagebox.showerror("ZIP inválido", f"Não é um ZIP válido:\n{zip_path}")
            return
        if not pdf_names:
            messagebox.showwarning("Nenhum PDF", f"ZIP sem PDFs:\n{zip_path}")
            return
        tmp_dir = tempfile.mkdtemp(prefix="onr_open_")
        self._temp_dirs.append(tmp_dir)
        try:
            with zipfile.ZipFile(zip_path, "r") as zf:
                for name in pdf_names:
                    zf.extract(name, tmp_dir)
        except Exception as exc:
            messagebox.showerror("Erro ao extrair ZIP", str(exc))
            return
        if len(pdf_names) == 1:
            try:
                os.startfile(os.path.join(tmp_dir, pdf_names[0]))
            except Exception as exc:
                messagebox.showerror("Erro ao abrir PDF", str(exc))
            return
        self._show_zip_open_dialog(zip_path, tmp_dir, pdf_names)

    def _show_zip_open_dialog(self, zip_path: str, tmp_dir: str, pdf_names: list) -> None:
        """Diálogo para escolher qual PDF do ZIP abrir."""
        dialog = tk.Toplevel(self)
        dialog.title("Selecionar PDF para Abrir")
        dialog.geometry("560x400")
        dialog.resizable(True, True)
        dialog.configure(bg=C_BG)
        dialog.grab_set()
        dialog.update_idletasks()
        px = self.winfo_x() + (self.winfo_width()  - 560) // 2
        py = self.winfo_y() + (self.winfo_height() - 400) // 2
        dialog.geometry(f"560x400+{px}+{py}")
        hdr = tk.Frame(dialog, bg=C_HEADER)
        hdr.pack(fill=tk.X)
        tk.Label(hdr, text="📂  Abrir PDF do ZIP",
                 font=("Segoe UI", 13, "bold"), bg=C_HEADER, fg="white", pady=10).pack()
        tk.Label(hdr, text=f"ZIP: {os.path.basename(zip_path)}",
                 font=("Segoe UI", 8), bg=C_HEADER, fg="#90CAF9").pack(pady=(0, 4))
        tk.Frame(hdr, bg="#0D47A1", height=2).pack(fill=tk.X)
        tk.Label(dialog, text="Selecione o PDF que deseja abrir:",
                 font=("Segoe UI", 9), bg=C_BG, fg=C_MUTED, anchor=tk.W
                 ).pack(fill=tk.X, padx=14, pady=(10, 4))
        list_frame = tk.Frame(dialog, bg=C_CARD, relief="solid", bd=1)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=14, pady=(0, 8))
        lb_scroll = ttk.Scrollbar(list_frame, orient=tk.VERTICAL)
        listbox = tk.Listbox(list_frame, font=("Segoe UI", 10),
                             yscrollcommand=lb_scroll.set,
                             selectmode=tk.SINGLE,
                             bg=C_CARD, fg="#212121",
                             selectbackground="#BBDEFB", selectforeground="#0D47A1",
                             relief="flat", bd=0)
        lb_scroll.config(command=listbox.yview)
        lb_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=4, pady=4)
        for name in pdf_names:
            listbox.insert(tk.END, f"  {os.path.basename(name)}")
        if pdf_names:
            listbox.selection_set(0)
        btn_bar = tk.Frame(dialog, bg=C_BG)
        btn_bar.pack(fill=tk.X, padx=14, pady=(0, 12))

        def do_open():
            sel_idx = listbox.curselection()
            if not sel_idx:
                messagebox.showwarning("Nenhum selecionado", "Selecione um PDF.", parent=dialog)
                return
            chosen = pdf_names[sel_idx[0]]
            dialog.destroy()
            try:
                os.startfile(os.path.join(tmp_dir, chosen))
            except Exception as exc:
                messagebox.showerror("Erro ao abrir PDF", str(exc))

        _make_btn(btn_bar, "📂  Abrir selecionado", do_open, C_SUCCESS).pack(side=tk.LEFT, padx=(0, 8))
        _make_btn(btn_bar, "Cancelar", dialog.destroy, C_MUTED, size=9).pack(side=tk.RIGHT)
        listbox.bind("<Double-1>", lambda _e: do_open())

    def _on_download_from_drive(self) -> None:
        sel = self._tree.selection()
        drive_sel = [iid for iid in sel if iid in self._drive_files]
        if not drive_sel:
            return
        local_folder = self._folder_var.get().strip()
        if not local_folder or not os.path.isdir(local_folder):
            messagebox.showerror("Pasta local não definida",
                                 "Selecione uma pasta local válida antes de baixar.")
            return
        items = [(iid, self._found_files[iid]) for iid in drive_sel]
        nomes = "\n".join(f"  • {os.path.basename(p)}" for _, p in items)
        ok = messagebox.askyesno("Confirmar download",
                                 f"Copiar {len(items)} arquivo(s) do OneDrive para:\n"
                                 f"{local_folder}\n\n{nomes}")
        if not ok:
            return
        erros, copiados = [], []
        for iid, src_path in items:
            fname = os.path.basename(src_path)
            dst_path = os.path.join(local_folder, fname)
            if os.path.exists(dst_path):
                resp = messagebox.askyesno("Arquivo já existe",
                                           f"O arquivo já existe na pasta local:\n{fname}\n\nSobrescrever?")
                if not resp:
                    continue
            try:
                shutil.copy2(src_path, dst_path)
                copiados.append((iid, fname, dst_path))
            except Exception as exc:
                erros.append(f"{fname}: {exc}")
        for iid, fname, dst_path in copiados:
            self._tree.item(iid,
                            values=("✅  Local", self._tree.item(iid, "values")[1], fname, dst_path),
                            tags=("found",))
            self._found_files[iid] = dst_path
            self._drive_files.discard(iid)
        if erros:
            messagebox.showerror("Erros ao copiar",
                                 "Alguns arquivos não puderam ser copiados:\n\n" + "\n".join(erros))
        elif copiados:
            messagebox.showinfo("Download concluído",
                                f"{len(copiados)} arquivo(s) copiado(s) para:\n{local_folder}")
            self._status_var.set(
                f"☁→📁  {len(copiados)} arquivo(s) baixado(s) do OneDrive para a pasta local.")
        self._recount()

    def _on_clear(self) -> None:
        self._ph_clear()
        self._proto_txt.delete("1.0", tk.END)
        self._ph_restore()
        self._tree.delete(*self._tree.get_children())
        self._found_files.clear()
        self._drive_files.clear()
        self._btn_open.config(state="disabled")
        self._btn_print.config(state="disabled")
        self._btn_download.config(state="disabled", text="⬇   BAIXAR DO DRIVE")
        self._update_counters(0, 0, 0, 0)
        self._status_var.set("Campos limpos. Pronto para nova consulta.")

    # ── Impressão ─────────────────────────────────────────────
    def _on_print_file(self) -> None:
        sel = self._tree.selection()
        if not sel:
            return
        found_sel = [iid for iid in sel if iid in self._found_files]
        if not found_sel:
            messagebox.showinfo("Aviso", "Selecione ao menos um item ✅/☁ para imprimir.")
            return
        paths, missing = [], []
        for iid in found_sel:
            p = self._found_files[iid]
            (paths if os.path.exists(p) else missing).append(p)
        if missing:
            messagebox.showwarning("Arquivos não encontrados",
                                   "Ignorados (não localizados no disco):\n" + "\n".join(missing))
        if not paths:
            return
        if len(paths) > 1:
            nomes = "\n".join(f"  • {os.path.basename(p)}" for p in paths)
            if not messagebox.askyesno("Confirmar impressão",
                                       f"Imprimir {len(paths)} arquivo(s)?\n\n{nomes}"):
                return
        for path in paths:
            if os.path.splitext(path)[1].lower() == ".zip":
                self._print_from_zip(path)
            else:
                self._print_file_direct(path)

    def _print_file_direct(self, path: str) -> None:
        fname = os.path.basename(path)
        ext   = os.path.splitext(path)[1].lower()
        if ext == ".pdf":
            ok, method = self._try_print_pdf(path)
        else:
            ok, method = self._shellexec_print(path), "ShellExecute"
        if ok:
            self._status_var.set(f"🖨  Enviado para impressão via {method}: {fname}")
        else:
            if messagebox.askyesno("Impressão automática indisponível",
                                   f"Não foi possível enviar diretamente.\n\nAbrir para imprimir manualmente?\n{path}"):
                try:
                    os.startfile(path)
                except Exception as exc:
                    messagebox.showerror("Erro ao abrir", str(exc))

    def _try_print_pdf(self, path: str) -> tuple[bool, str]:
        sumatra = self._find_exe([
            r"C:\Program Files\SumatraPDF\SumatraPDF.exe",
            r"C:\Program Files (x86)\SumatraPDF\SumatraPDF.exe",
        ], which="SumatraPDF")
        if sumatra:
            try:
                subprocess.Popen([sumatra, "-print-to-default", "-silent", path],
                                 stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                return True, "SumatraPDF"
            except Exception:
                pass
        acrobat = self._find_exe([
            r"C:\Program Files\Adobe\Acrobat DC\Acrobat\Acrobat.exe",
            r"C:\Program Files (x86)\Adobe\Acrobat DC\Acrobat\Acrobat.exe",
            r"C:\Program Files\Adobe\Acrobat Reader DC\Reader\AcroRd32.exe",
            r"C:\Program Files (x86)\Adobe\Acrobat Reader DC\Reader\AcroRd32.exe",
        ], which="AcroRd32")
        if acrobat:
            try:
                subprocess.Popen([acrobat, "/p", "/h", path],
                                 stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                return True, "Adobe Reader"
            except Exception:
                pass
        foxit = self._find_exe([
            r"C:\Program Files\Foxit Software\Foxit PDF Reader\FoxitPDFReader.exe",
            r"C:\Program Files (x86)\Foxit Software\Foxit Reader\FoxitReader.exe",
        ], which="FoxitReader")
        if foxit:
            try:
                subprocess.Popen([foxit, "/p", path],
                                 stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                return True, "Foxit Reader"
            except Exception:
                pass
        default_printer = self._get_default_printer()
        if default_printer:
            try:
                import ctypes
                ret = ctypes.windll.shell32.ShellExecuteW(
                    None, "printto", path, f'"{default_printer}"', None, 0)
                if ret > 32:
                    return True, f"impressora padrão ({default_printer})"
            except Exception:
                pass
        if self._shellexec_print(path):
            return True, "ShellExecute"
        return False, ""

    @staticmethod
    def _find_exe(paths: list, which: str | None = None) -> str | None:
        for p in paths:
            if os.path.isfile(p):
                return p
        if which:
            found = shutil.which(which)
            if found:
                return found
        return None

    @staticmethod
    def _get_default_printer() -> str | None:
        try:
            import win32print
            return win32print.GetDefaultPrinter()
        except ImportError:
            pass
        try:
            import winreg
            key = winreg.OpenKey(winreg.HKEY_CURRENT_USER,
                                 r"Software\Microsoft\Windows NT\CurrentVersion\Windows")
            printer, _ = winreg.QueryValueEx(key, "Device")
            winreg.CloseKey(key)
            return printer.split(",")[0].strip()
        except Exception:
            return None

    @staticmethod
    def _shellexec_print(path: str) -> bool:
        try:
            import ctypes
            ret = ctypes.windll.shell32.ShellExecuteW(None, "print", path, None, None, 0)
            return ret > 32
        except Exception:
            return False

    def _print_from_zip(self, zip_path: str) -> None:
        try:
            with zipfile.ZipFile(zip_path, "r") as zf:
                pdf_names = [n for n in zf.namelist()
                             if n.lower().endswith(".pdf") and not n.startswith("__MACOSX")]
        except zipfile.BadZipFile:
            messagebox.showerror("ZIP inválido", f"Não é um ZIP válido:\n{zip_path}")
            return
        if not pdf_names:
            messagebox.showwarning("Nenhum PDF", f"ZIP sem PDFs:\n{zip_path}")
            return
        tmp_dir = tempfile.mkdtemp(prefix="onr_print_")
        self._temp_dirs.append(tmp_dir)
        try:
            with zipfile.ZipFile(zip_path, "r") as zf:
                for name in pdf_names:
                    zf.extract(name, tmp_dir)
        except Exception as exc:
            messagebox.showerror("Erro ao extrair ZIP", str(exc))
            return
        if len(pdf_names) == 1:
            self._print_file_direct(os.path.join(tmp_dir, pdf_names[0]))
            return
        self._show_zip_print_dialog(zip_path, tmp_dir, pdf_names)

    def _show_zip_print_dialog(self, zip_path: str, tmp_dir: str, pdf_names: list) -> None:
        dialog = tk.Toplevel(self)
        dialog.title("Selecionar PDFs para Imprimir")
        dialog.geometry("560x400")
        dialog.resizable(True, True)
        dialog.configure(bg=C_BG)
        dialog.grab_set()
        dialog.update_idletasks()
        px = self.winfo_x() + (self.winfo_width()  - 560) // 2
        py = self.winfo_y() + (self.winfo_height() - 400) // 2
        dialog.geometry(f"560x400+{px}+{py}")
        hdr = tk.Frame(dialog, bg=C_HEADER)
        hdr.pack(fill=tk.X)
        tk.Label(hdr, text="🖨  Imprimir PDFs do ZIP",
                 font=("Segoe UI", 13, "bold"), bg=C_HEADER, fg="white", pady=10).pack()
        tk.Label(hdr, text=f"ZIP: {os.path.basename(zip_path)}",
                 font=("Segoe UI", 8), bg=C_HEADER, fg="#90CAF9").pack(pady=(0, 4))
        tk.Frame(hdr, bg="#0D47A1", height=2).pack(fill=tk.X)
        tk.Label(dialog, text="Marque os PDFs que deseja imprimir:",
                 font=("Segoe UI", 9), bg=C_BG, fg=C_MUTED, anchor=tk.W
                 ).pack(fill=tk.X, padx=14, pady=(10, 4))
        list_frame = tk.Frame(dialog, bg=C_CARD, relief="solid", bd=1)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=14, pady=(0, 8))
        canvas = tk.Canvas(list_frame, bg=C_CARD, highlightthickness=0)
        scroll = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=canvas.yview)
        inner  = tk.Frame(canvas, bg=C_CARD)
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=inner, anchor="nw")
        canvas.configure(yscrollcommand=scroll.set)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        check_vars: list[tk.BooleanVar] = []
        for i, name in enumerate(pdf_names):
            var = tk.BooleanVar(value=True)
            check_vars.append(var)
            row_bg = C_CARD if i % 2 == 0 else "#F7F9FC"
            tk.Checkbutton(inner, text=f"  {os.path.basename(name)}", variable=var,
                           font=("Segoe UI", 10), bg=row_bg, fg="#212121",
                           activebackground=row_bg, anchor=tk.W,
                           ).pack(fill=tk.X, padx=8, pady=3, ipadx=4, ipady=3)
        btn_bar = tk.Frame(dialog, bg=C_BG)
        btn_bar.pack(fill=tk.X, padx=14, pady=(0, 12))

        def do_print():
            selected = [os.path.join(tmp_dir, pdf_names[i])
                        for i, v in enumerate(check_vars) if v.get()]
            if not selected:
                messagebox.showwarning("Nenhum selecionado", "Marque ao menos um PDF.", parent=dialog)
                return
            dialog.destroy()
            for p in selected:
                self._print_file_direct(p)

        _make_btn(btn_bar, "🖨  Imprimir selecionados", do_print, C_PRINT).pack(side=tk.LEFT, padx=(0, 8))
        _make_btn(btn_bar, "✔  Todos",  lambda: [v.set(True)  for v in check_vars], C_ACCENT, size=9).pack(side=tk.LEFT, padx=(0, 4))
        _make_btn(btn_bar, "✖  Nenhum", lambda: [v.set(False) for v in check_vars], C_MUTED,  size=9).pack(side=tk.LEFT)
        _make_btn(btn_bar, "Cancelar",  dialog.destroy, C_MUTED, size=9).pack(side=tk.RIGHT)

    # ── Utilitários ───────────────────────────────────────────
    def _get_protocols(self) -> list:
        if self._ph_active:
            return []
        return [line.strip()
                for line in self._proto_txt.get("1.0", tk.END).splitlines()
                if line.strip()]

    def _update_counters(self, total, local, drive, not_found) -> None:
        self._lbl_total.config(text=f"Total: {total}")
        self._lbl_found.config(text=f"✅  Local: {local}")
        self._lbl_drive.config(text=f"☁  Drive: {drive}")
        self._lbl_notfound.config(text=f"❌  Não encontrados: {not_found}")

    def _recount(self) -> None:
        local = drive = notfound = total = 0
        for iid in self._tree.get_children():
            total += 1
            tags = self._tree.item(iid, "tags")
            if "found"    in tags: local    += 1
            elif "drive"  in tags: drive    += 1
            elif "notfound" in tags: notfound += 1
        self._update_counters(total, local, drive, notfound)

    def _center_window(self) -> None:
        self.update_idletasks()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        w,  h  = self.winfo_width(), self.winfo_height()
        self.geometry(f"{w}x{h}+{(sw - w) // 2}+{(sh - h) // 2}")

    def _on_close(self) -> None:
        for tmp in self._temp_dirs:
            shutil.rmtree(tmp, ignore_errors=True)
        self.destroy()


# ══════════════════════════════════════════════════════════════
#  SISTEMA DE ATUALIZAÇÃO AUTOMÁTICA
# ══════════════════════════════════════════════════════════════
class UpdateManager:
    """Gerencia verificação e download de atualizações do aplicativo."""

    @staticmethod
    def check_local_version() -> str:
        """Verifica versão local lendo arquivo version.txt na pasta do executável."""
        try:
            if getattr(sys, 'frozen', False):
                # Rodando como EXE
                app_dir = os.path.dirname(sys.executable)
            else:
                # Rodando como script Python
                app_dir = os.path.dirname(os.path.abspath(__file__))
            
            version_file = os.path.join(app_dir, "version.txt")
            if os.path.exists(version_file):
                with open(version_file, "r", encoding="utf-8") as f:
                    return f.read().strip()
        except Exception:
            pass
        return APP_VERSION

    @staticmethod
    def check_remote_version() -> str | None:
        """Verifica versão remota no OneDrive/pasta compartilhada."""
        try:
            # Tenta ler arquivo de versão na pasta de atualização local (OneDrive sincronizado)
            if os.path.exists(UPDATE_FOLDER):
                version_file = os.path.join(UPDATE_FOLDER, "version.txt")
                if os.path.exists(version_file):
                    with open(version_file, "r", encoding="utf-8") as f:
                        return f.read().strip()
        except Exception:
            pass
        return None

    @staticmethod
    def has_update() -> tuple[bool, str, str]:
        """Verifica se há atualização disponível. Retorna (has_update, local_ver, remote_ver)."""
        local = UpdateManager.check_local_version()
        remote = UpdateManager.check_remote_version()
        
        if not remote:
            return False, local, local
        
        # Compara versões (simples comparação de strings)
        has_update = remote != local
        return has_update, local, remote

    @staticmethod
    def download_update() -> str | None:
        """Baixa o EXE atualizado para pasta temporária. Retorna caminho do arquivo baixado."""
        try:
            if os.path.exists(UPDATE_FOLDER):
                exe_file = os.path.join(UPDATE_FOLDER, "BrejeiroCertidoes.exe")
                if os.path.exists(exe_file):
                    # Copia para temp
                    temp_dir = tempfile.mkdtemp(prefix="onr_update_")
                    temp_exe = os.path.join(temp_dir, "BrejeiroCertidoes_new.exe")
                    shutil.copy2(exe_file, temp_exe)
                    return temp_exe
        except Exception as exc:
            print(f"Erro ao baixar atualização: {exc}")
        return None

    @staticmethod
    def apply_update(new_exe_path: str) -> bool:
        """Aplica a atualização substituindo o executável atual."""
        try:
            if getattr(sys, 'frozen', False):
                current_exe = sys.executable
                backup_exe = current_exe + ".old"
                
                # Backup do atual
                if os.path.exists(backup_exe):
                    os.remove(backup_exe)
                shutil.move(current_exe, backup_exe)
                
                # Move novo para lugar
                shutil.move(new_exe_path, current_exe)
                
                # Remove backup
                if os.path.exists(backup_exe):
                    os.remove(backup_exe)
                
                return True
        except Exception as exc:
            print(f"Erro ao aplicar atualização: {exc}")
        return False


# ══════════════════════════════════════════════════════════════
#  GERENCIAMENTO DE USUÁRIOS (CLIENTE API)
# ══════════════════════════════════════════════════════════════
class UserManager:
    """Gerencia usuários e autenticação via API REST centralizada."""

    def __init__(self, server_url: str = None):
        # URL do servidor de autenticação
        if server_url:
            self._server_url = server_url.rstrip('/')
        else:
            # Tenta ler do arquivo de configuração
            self._server_url = self._load_server_url()
        self._connected = False
        self._check_connection()
        # Token secreto para operações de admin (deve ser o mesmo do servidor)
        self._admin_token = 'brejeiro2026_seguro_9283_XY77'

    def _load_server_url(self) -> str:
        """Carrega URL do servidor do arquivo de configuração."""
        config_file = self._get_config_file()
        try:
            if os.path.exists(config_file):
                with open(config_file, "r", encoding="utf-8") as f:
                    config = json.load(f)
                    return config.get("server_url", "https://brejeiro.onrender.com")
        except Exception:
            pass
        return "https://brejeiro.onrender.com"

    def _get_config_file(self) -> str:
        """Retorna o caminho do arquivo de configuração."""
        if getattr(sys, 'frozen', False):
            app_dir = os.path.dirname(sys.executable)
        else:
            app_dir = os.path.dirname(os.path.abspath(__file__))
        return os.path.join(app_dir, "auth_config.json")

    def _check_connection(self) -> bool:
        """Verifica se consegue conectar ao servidor."""
        try:
            url = f"{self._server_url}/health"
            with urllib.request.urlopen(url, timeout=5) as response:
                self._connected = response.status == 200
                return self._connected
        except Exception:
            self._connected = False
            return False

    def _api_request(self, method: str, endpoint: str, data: dict = None, is_admin: bool = False) -> tuple[bool, dict]:
        """Faz uma requisição à API do servidor."""
        url = f"{self._server_url}/api/{endpoint}"
        
        try:
            headers = {'Content-Type': 'application/json'}
            # Adiciona token de admin para rotas protegidas
            if is_admin:
                headers['X-Admin-Token'] = self._admin_token
            
            if method == "GET":
                req = urllib.request.Request(url, headers=headers, method='GET')
                with urllib.request.urlopen(req, timeout=10) as response:
                    return True, json.loads(response.read().decode('utf-8'))
            
            elif method in ["POST", "PUT", "DELETE"]:
                json_data = json.dumps(data).encode('utf-8')
                req = urllib.request.Request(
                    url,
                    data=json_data,
                    headers=headers,
                    method=method
                )
                with urllib.request.urlopen(req, timeout=10) as response:
                    return True, json.loads(response.read().decode('utf-8'))
        except urllib.error.HTTPError as e:
            try:
                error_data = json.loads(e.read().decode('utf-8'))
                return False, error_data
            except:
                return False, {"message": f"Erro HTTP {e.code}"}
        except urllib.error.URLError:
            return False, {"message": "Não foi possível conectar ao servidor de autenticação."}
        except Exception as e:
            return False, {"message": f"Erro: {str(e)}"}
        
        return False, {"message": "Erro desconhecido"}

    def is_connected(self) -> bool:
        """Retorna se está conectado ao servidor."""
        return self._connected

    def get_server_url(self) -> str:
        """Retorna a URL do servidor configurado."""
        return self._server_url

    def set_server_url(self, url: str) -> bool:
        """Configura a URL do servidor e salva no arquivo de configuração."""
        self._server_url = url.rstrip('/')
        config_file = self._get_config_file()
        try:
            with open(config_file, "w", encoding="utf-8") as f:
                json.dump({"server_url": self._server_url}, f, indent=2)
            self._check_connection()
            return True
        except Exception:
            return False

    def authenticate(self, username: str, password: str) -> tuple[bool, dict]:
        """Autentica um usuário via API. Retorna (sucesso, dados_do_usuário)."""
        success, response = self._api_request("POST", "auth/login", {
            "username": username,
            "password": password
        })
        
        if success and response.get("success"):
            user_data = response.get("user", {})
            return True, user_data
        
        return False, {}

    def get_all_users(self) -> list[tuple[str, dict]]:
        """Retorna lista de todos os usuários via API."""
        success, response = self._api_request("GET", "users", is_admin=True)
        
        if success and response.get("success"):
            users_list = []
            for user in response.get("users", []):
                users_list.append((user["username"], {
                    "is_admin": user.get("is_admin", False)
                }))
            return users_list
        
        return []

    def add_user(self, username: str, password: str, is_admin: bool = False) -> tuple[bool, str]:
        """Adiciona um novo usuário via API. Retorna (sucesso, mensagem)."""
        if not username or not password:
            return False, "Usuário e senha são obrigatórios."
        
        success, response = self._api_request("POST", "users", {
            "username": username,
            "password": password,
            "is_admin": is_admin
        }, is_admin=True)
        
        if success:
            return True, response.get("message", "Usuário criado com sucesso.")
        return False, response.get("message", "Erro ao criar usuário.")

    def update_user(self, username: str, password: str = None, is_admin: bool = None) -> tuple[bool, str]:
        """Atualiza um usuário existente via API. Retorna (sucesso, mensagem)."""
        if username == "Comprasoja" and is_admin is False:
            return False, "Não é possível remover privilégios de admin do usuário principal."
        
        data = {}
        if password is not None:
            data["password"] = password
        if is_admin is not None:
            data["is_admin"] = is_admin
        
        success, response = self._api_request("PUT", f"users/{username}", data, is_admin=True)
        
        if success:
            return True, response.get("message", "Usuário atualizado com sucesso.")
        return False, response.get("message", "Erro ao atualizar usuário.")

    def delete_user(self, username: str) -> tuple[bool, str]:
        """Remove um usuário via API. Retorna (sucesso, mensagem)."""
        if username == "Comprasoja":
            return False, "Não é possível remover o usuário admin principal."
        
        success, response = self._api_request("DELETE", f"users/{username}", is_admin=True)
        
        if success:
            return True, response.get("message", "Usuário removido com sucesso.")
        return False, response.get("message", "Erro ao remover usuário.")


# ══════════════════════════════════════════════════════════════
#  PAINEL DE ADMINISTRAÇÃO
# ══════════════════════════════════════════════════════════════
class AdminPanel(tk.Toplevel):
    """Painel de administração para gerenciar usuários."""

    def __init__(self, parent: tk.Widget, user_manager: UserManager, current_user: str) -> None:
        super().__init__(parent)
        self.title("Painel de Administração")
        self.geometry("700x550")
        self.resizable(True, True)
        self.configure(bg=C_BG)
        self.grab_set()
        self._user_manager = user_manager
        self._current_user = current_user

        try:
            from ctypes import windll
            windll.shcore.SetProcessDpiAwareness(1)
        except Exception:
            pass

        try:
            self.iconbitmap("icon.ico")
        except Exception:
            pass

        self._center_window(parent)
        self._build_ui()
        self._refresh_users_list()

    def _center_window(self, parent: tk.Widget) -> None:
        self.update_idletasks()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        w, h = 700, 550
        x = (sw - w) // 2
        y = (sh - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")

    def _build_ui(self) -> None:
        # Header
        hdr = tk.Frame(self, bg=C_HEADER)
        hdr.pack(fill=tk.X)
        tk.Label(hdr, text="👤  Gerenciamento de Usuários",
                 font=("Segoe UI", 14, "bold"), bg=C_HEADER, fg="white", pady=12).pack()
        tk.Frame(hdr, bg="#0D47A1", height=3).pack(fill=tk.X, pady=(0, 0))

        # Container principal
        container = tk.Frame(self, bg=C_BG, padx=20, pady=15)
        container.pack(fill=tk.BOTH, expand=True)

        # Lista de usuários
        list_frame = tk.LabelFrame(
            container, text=" Usuários Cadastrados ",
            font=("Segoe UI", 9, "bold"),
            bg=C_CARD, fg=C_ACCENT,
            relief="solid", bd=1,
            padx=10, pady=8,
        )
        list_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

        # Treeview para usuários
        cols = ("username", "is_admin")
        self._users_tree = ttk.Treeview(list_frame, columns=cols, show="headings", selectmode="browse")
        self._users_tree.heading("username", text="Usuário", anchor=tk.W)
        self._users_tree.heading("is_admin", text="Admin", anchor=tk.CENTER)
        self._users_tree.column("username", width=250, minwidth=150)
        self._users_tree.column("is_admin", width=100, minwidth=80, anchor=tk.CENTER)
        self._users_tree.tag_configure("admin", foreground=C_ACCENT)
        self._users_tree.tag_configure("normal", foreground="#212121")
        
        sy = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self._users_tree.yview)
        sx = ttk.Scrollbar(list_frame, orient=tk.HORIZONTAL, command=self._users_tree.xview)
        self._users_tree.configure(yscrollcommand=sy.set, xscrollcommand=sx.set)
        sy.pack(side=tk.RIGHT, fill=tk.Y)
        sx.pack(side=tk.BOTTOM, fill=tk.X)
        self._users_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self._users_tree.bind("<<TreeviewSelect>>", self._on_user_select)

        # Botões de ação
        btn_frame = tk.Frame(container, bg=C_BG)
        btn_frame.pack(fill=tk.X, pady=(0, 10))

        _make_btn(btn_frame, "➕  Novo Usuário", self._on_add_user, C_SUCCESS, size=9
                  ).pack(side=tk.LEFT, padx=(0, 6))
        _make_btn(btn_frame, "✏️  Editar", self._on_edit_user, C_ACCENT, size=9
                  ).pack(side=tk.LEFT, padx=(0, 6))
        _make_btn(btn_frame, "🗑️  Excluir", self._on_delete_user, C_DANGER, size=9
                  ).pack(side=tk.LEFT, padx=(0, 6))
        _make_btn(btn_frame, "🔄  Atualizar", self._refresh_users_list, C_MUTED, size=9
                  ).pack(side=tk.LEFT)
        _make_btn(btn_frame, "Fechar", self.destroy, C_MUTED, size=9
                  ).pack(side=tk.RIGHT)

        # Botão desabilitado inicialmente
        self._btn_edit = btn_frame.winfo_children()[1]
        self._btn_delete = btn_frame.winfo_children()[2]
        self._btn_edit.config(state="disabled")
        self._btn_delete.config(state="disabled")

    def _refresh_users_list(self) -> None:
        """Atualiza a lista de usuários no treeview."""
        self._users_tree.delete(*self._users_tree.get_children())
        for username, data in self._user_manager.get_all_users():
            is_admin = "Sim" if data.get("is_admin", False) else "Não"
            tag = "admin" if data.get("is_admin", False) else "normal"
            self._users_tree.insert("", tk.END, values=(username, is_admin), tags=(tag,))
        self._btn_edit.config(state="disabled")
        self._btn_delete.config(state="disabled")

    def _on_user_select(self, _event=None) -> None:
        """Habilita botões quando um usuário é selecionado."""
        sel = self._users_tree.selection()
        if sel:
            username = self._users_tree.item(sel[0], "values")[0]
            self._btn_edit.config(state="normal")
            # Não permite excluir o próprio usuário ou o admin
            if username == "admin" or username == self._current_user:
                self._btn_delete.config(state="disabled")
            else:
                self._btn_delete.config(state="normal")
        else:
            self._btn_edit.config(state="disabled")
            self._btn_delete.config(state="disabled")

    def _on_add_user(self) -> None:
        """Abre diálogo para adicionar novo usuário."""
        self._show_user_dialog()

    def _on_edit_user(self) -> None:
        """Abre diálogo para editar usuário selecionado."""
        sel = self._users_tree.selection()
        if not sel:
            return
        username = self._users_tree.item(sel[0], "values")[0]
        user_data = self._user_manager._users[username]
        self._show_user_dialog(username, user_data)

    def _on_delete_user(self) -> None:
        """Remove o usuário selecionado."""
        sel = self._users_tree.selection()
        if not sel:
            return
        username = self._users_tree.item(sel[0], "values")[0]
        if messagebox.askyesno("Confirmar Exclusão",
                               f"Deseja realmente remover o usuário '{username}'?"):
            success, msg = self._user_manager.delete_user(username)
            if success:
                messagebox.showinfo("Sucesso", msg)
                self._refresh_users_list()
            else:
                messagebox.showerror("Erro", msg)

    def _show_user_dialog(self, username: str = None, user_data: dict = None) -> None:
        """Mostra diálogo para criar/editar usuário."""
        dialog = tk.Toplevel(self)
        dialog.title("Novo Usuário" if username is None else "Editar Usuário")
        dialog.geometry("400x320")
        dialog.resizable(False, False)
        dialog.configure(bg=C_BG)
        dialog.grab_set()
        self._center_dialog(dialog)

        # Header
        hdr = tk.Frame(dialog, bg=C_HEADER)
        hdr.pack(fill=tk.X)
        title = "Editar Usuário" if username else "Novo Usuário"
        tk.Label(hdr, text=f"👤  {title}",
                 font=("Segoe UI", 13, "bold"), bg=C_HEADER, fg="white", pady=10).pack()
        tk.Frame(hdr, bg="#0D47A1", height=2).pack(fill=tk.X)

        # Container
        container = tk.Frame(dialog, bg=C_BG, padx=25, pady=20)
        container.pack(fill=tk.BOTH, expand=True)

        # Campo usuário
        tk.Label(container, text="Usuário:", font=("Segoe UI", 10, "bold"),
                 bg=C_BG, fg="#333").pack(anchor=tk.W, pady=(0, 5))
        user_var = tk.StringVar(value=username or "")
        user_entry = tk.Entry(container, textvariable=user_var, font=("Segoe UI", 11),
                              relief="solid", bd=1, width=30)
        user_entry.pack(ipady=8, pady=(0, 15))
        if username:
            user_entry.config(state="readonly")

        # Campo senha
        tk.Label(container, text="Senha:", font=("Segoe UI", 10, "bold"),
                 bg=C_BG, fg="#333").pack(anchor=tk.W, pady=(0, 5))
        pass_var = tk.StringVar(value=user_data["password"] if user_data else "")
        pass_entry = tk.Entry(container, textvariable=pass_var, font=("Segoe UI", 11),
                              relief="solid", bd=1, width=30, show="•")
        pass_entry.pack(ipady=8, pady=(0, 15))

        # Checkbox admin
        admin_var = tk.BooleanVar(value=user_data.get("is_admin", False) if user_data else False)
        admin_check = tk.Checkbutton(container, text="Usuário Administrador",
                                     variable=admin_var, font=("Segoe UI", 10),
                                     bg=C_BG, fg="#333", activebackground=C_BG)
        admin_check.pack(anchor=tk.W, pady=(0, 20))

        # Botões
        btn_frame = tk.Frame(container, bg=C_BG)
        btn_frame.pack(fill=tk.X)

        def on_save():
            new_username = user_var.get().strip()
            new_password = pass_var.get().strip()
            is_admin = admin_var.get()

            # Debug para verificar valores
            print(f"DEBUG - Username: '{new_username}', Password: '{new_password}', len: {len(new_password)}")

            if not new_username:
                messagebox.showwarning("Campo obrigatório",
                                       "O campo Usuário é obrigatório.", parent=dialog)
                user_entry.focus_set()
                return
            
            if not new_password:
                messagebox.showwarning("Campo obrigatório",
                                       "O campo Senha é obrigatório.", parent=dialog)
                pass_entry.focus_set()
                return

            if username:
                # Editar usuário existente
                success, msg = self._user_manager.update_user(
                    username, password=new_password, is_admin=is_admin
                )
            else:
                # Criar novo usuário
                success, msg = self._user_manager.add_user(
                    new_username, new_password, is_admin
                )

            if success:
                messagebox.showinfo("Sucesso", msg, parent=dialog)
                dialog.destroy()
                self._refresh_users_list()
            else:
                messagebox.showerror("Erro", msg, parent=dialog)

        _make_btn(btn_frame, "💾  Salvar", on_save, C_SUCCESS, size=10
                  ).pack(side=tk.LEFT, expand=True, fill=tk.X, padx=(0, 8))
        _make_btn(btn_frame, "Cancelar", dialog.destroy, C_MUTED, size=10
                  ).pack(side=tk.LEFT, expand=True, fill=tk.X)

        if username is None:
            user_entry.focus_set()
        else:
            pass_entry.focus_set()

    def _center_dialog(self, dialog: tk.Toplevel) -> None:
        dialog.update_idletasks()
        px = self.winfo_x() + (self.winfo_width() - 400) // 2
        py = self.winfo_y() + (self.winfo_height() - 320) // 2
        dialog.geometry(f"400x320+{px}+{py}")


# ══════════════════════════════════════════════════════════════
#  TELA DE LOGIN
# ══════════════════════════════════════════════════════════════
class LoginDialog(tk.Toplevel):
    """Janela de login para autenticação antes de acessar o sistema."""

    def __init__(self, parent: tk.Widget, user_manager: UserManager) -> None:
        super().__init__(parent)
        self.title("Login — Certidões ONR")
        self.geometry("420x320")
        self.resizable(False, False)
        self.configure(bg=C_BG)
        self.grab_set()
        self._authenticated = False
        self._user_manager = user_manager
        self._authenticated_user = None
        self._user_data = None

        try:
            from ctypes import windll
            windll.shcore.SetProcessDpiAwareness(1)
        except Exception:
            pass

        try:
            self.iconbitmap("icon.ico")
        except Exception:
            pass

        self._center_window(parent)
        self._build_ui()

        self.protocol("WM_DELETE_WINDOW", self._on_close)

        # Foco no campo usuário
        self.after(100, lambda: self._user_entry.focus_set())

    def _center_window(self, parent: tk.Widget) -> None:
        self.update_idletasks()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        w, h = 420, 320
        x = (sw - w) // 2
        y = (sh - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")

    def _build_ui(self) -> None:
        # Header
        hdr = tk.Frame(self, bg=C_HEADER)
        hdr.pack(fill=tk.X)
        tk.Label(hdr, text="🔐  Acesso ao Sistema",
                 font=("Segoe UI", 14, "bold"), bg=C_HEADER, fg="white", pady=12).pack()
        tk.Frame(hdr, bg="#0D47A1", height=3).pack(fill=tk.X, pady=(0, 0))

        # Container principal
        container = tk.Frame(self, bg=C_BG, padx=30, pady=25)
        container.pack(fill=tk.BOTH, expand=True)

        # Campo usuário
        tk.Label(container, text="Usuário:", font=("Segoe UI", 10, "bold"),
                 bg=C_BG, fg="#333").pack(anchor=tk.W, pady=(0, 5))
        self._user_var = tk.StringVar()
        self._user_entry = tk.Entry(
            container,
            textvariable=self._user_var,
            font=("Segoe UI", 11),
            relief="solid", bd=1,
            width=30,
        )
        self._user_entry.pack(ipady=8, pady=(0, 15))
        self._user_entry.bind("<Return>", lambda e: self._pass_entry.focus_set())

        # Campo senha
        tk.Label(container, text="Senha:", font=("Segoe UI", 10, "bold"),
                 bg=C_BG, fg="#333").pack(anchor=tk.W, pady=(0, 5))
        self._pass_var = tk.StringVar()
        self._pass_entry = tk.Entry(
            container,
            textvariable=self._pass_var,
            font=("Segoe UI", 11),
            relief="solid", bd=1,
            width=30,
            show="•",
        )
        self._pass_entry.pack(ipady=8, pady=(0, 20))
        self._pass_entry.bind("<Return>", lambda e: self._on_login())

        # Botões
        btn_frame = tk.Frame(container, bg=C_BG)
        btn_frame.pack(fill=tk.X, pady=(10, 0))

        _make_btn(btn_frame, "🔑  Entrar", self._on_login, C_ACCENT, size=10
                  ).pack(side=tk.LEFT, expand=True, fill=tk.X, padx=(0, 8))
        _make_btn(btn_frame, "Cancelar", self._on_close, C_MUTED, size=10
                  ).pack(side=tk.LEFT, expand=True, fill=tk.X)

        # Mensagem de erro
        self._error_label = tk.Label(
            container,
            text="",
            font=("Segoe UI", 9),
            bg=C_BG, fg=C_DANGER,
        )
        self._error_label.pack(pady=(15, 0))

    def _on_login(self) -> None:
        user = self._user_var.get().strip()
        password = self._pass_var.get().strip()

        success, user_data = self._user_manager.authenticate(user, password)
        if success:
            self._authenticated = True
            self._authenticated_user = user
            self._user_data = user_data
            self.destroy()
        else:
            self._error_label.config(text="Usuário ou senha incorretos.")
            self._pass_var.set("")
            self._pass_entry.focus_set()

    def _on_close(self) -> None:
        self._authenticated = False
        self.destroy()

    def is_authenticated(self) -> bool:
        return self._authenticated

    def get_authenticated_user(self) -> tuple[str, dict]:
        """Retorna (username, user_data) do usuário autenticado."""
        return self._authenticated_user, self._user_data


# ══════════════════════════════════════════════════════════════
#  MAIN APP — Sistema de abas (entry point principal)
# ══════════════════════════════════════════════════════════════
class MainApp(tk.Tk):
    """
    Janela principal com sistema de abas:
      - Aba 1: Consulta ONR (busca de arquivos local + OneDrive)
      - Aba 2: Planilha de Controle (integração com Excel)
    """

    _PLACEHOLDER = "P26070520844D\nP26070520845D\nP26070520846D"

    def __init__(self, user_manager: UserManager, current_user: str, user_data: dict) -> None:
        super().__init__()

        try:
            from ctypes import windll
            windll.shcore.SetProcessDpiAwareness(1)
        except Exception:
            pass

        self.title("Certidões ONR — Brejeiro")
        self.geometry("1100x820")
        self.minsize(860, 620)
        self.configure(bg=C_BG)

        try:
            self.iconbitmap("icon.ico")
        except Exception:
            pass

        # Estado da aba 1
        self._found_files: dict = {}
        self._drive_files: set  = set()
        self._ph_active: bool = True
        self._temp_dirs: list = []

        # Gerenciamento de usuários
        self._user_manager = user_manager
        self._current_user = current_user
        self._user_data = user_data

        self.protocol("WM_DELETE_WINDOW", self._on_close)

        self._apply_styles()
        self._build_ui()
        self._center_window()

    # ──────────────────────────────────────────────────────────
    #  ESTILOS
    # ──────────────────────────────────────────────────────────
    def _apply_styles(self) -> None:
        s = ttk.Style(self)
        s.theme_use("clam")

        s.configure("Treeview", font=("Segoe UI", 10), rowheight=30,
                    background=C_CARD, fieldbackground=C_CARD, foreground="#212121")
        s.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"),
                    background=C_ACCENT, foreground="white", relief="flat")
        s.map("Treeview.Heading", background=[("active", "#0D47A1")])
        s.map("Treeview",
              background=[("selected", "#BBDEFB")],
              foreground=[("selected", "#0D47A1")])

        # Estilo do Notebook
        s.configure("TNotebook", background=C_BG, borderwidth=0)
        s.configure("TNotebook.Tab",
                    font=("Segoe UI", 10, "bold"),
                    padding=[16, 8],
                    background="#D0D8E4",
                    foreground="#1B3A5C")
        s.map("TNotebook.Tab",
              background=[("selected", C_HEADER), ("active", "#B0BEC5")],
              foreground=[("selected", "white"),   ("active", "#1B3A5C")])

    # ──────────────────────────────────────────────────────────
    #  CONSTRUÇÃO DO LAYOUT PRINCIPAL
    # ──────────────────────────────────────────────────────────
    def _build_ui(self) -> None:
        self._build_header()

        # Barra de status compartilhada (rodapé)
        self._status_var = tk.StringVar(value="Pronto.")
        tk.Label(self, textvariable=self._status_var, font=("Segoe UI", 9),
                 bg=C_STATUS_BG, fg="#455A64", anchor=tk.W, padx=10, pady=4,
                 relief="sunken").pack(side=tk.BOTTOM, fill=tk.X)

        # Notebook
        self._notebook = ttk.Notebook(self)
        self._notebook.pack(fill=tk.BOTH, expand=True, padx=0, pady=0)

        # ── Aba 1: Consulta ONR ───────────────────────────────
        self._tab1 = tk.Frame(self._notebook, bg=C_BG)
        self._notebook.add(self._tab1, text="  🔍  Consulta ONR — Arquivos  ")
        self._build_tab1_content()

        # ── Aba 2: Planilha de Controle ───────────────────────
        self._tab2_outer = tk.Frame(self._notebook, bg=C_BG)
        self._notebook.add(self._tab2_outer, text="  📊  Planilha de Controle  ")
        self._build_tab2_content()

    # ── Cabeçalho ─────────────────────────────────────────────
    def _build_header(self) -> None:
        hdr = tk.Frame(self, bg=C_HEADER)
        hdr.pack(fill=tk.X)

        # Container para título e botão de usuário
        header_content = tk.Frame(hdr, bg=C_HEADER)
        header_content.pack(fill=tk.X, padx=20, pady=10)

        # Título
        title_frame = tk.Frame(header_content, bg=C_HEADER)
        title_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        tk.Label(title_frame, text="Certidões ONR — Brejeiro",
                 font=("Segoe UI", 17, "bold"), bg=C_HEADER, fg="white").pack(anchor=tk.W)
        tk.Label(title_frame, text="Consulta de arquivos + Planilha de Controle de Certidões de Penhor (Goiás)",
                 font=("Segoe UI", 9), bg=C_HEADER, fg="#90CAF9").pack(anchor=tk.W, pady=(2, 0))

        # Botão de usuário (👤)
        user_frame = tk.Frame(header_content, bg=C_HEADER)
        user_frame.pack(side=tk.RIGHT)
        
        # Indicador de conexão
        conn_color = C_SUCCESS if self._user_manager.is_connected() else C_DANGER
        conn_text = "🟢 Online" if self._user_manager.is_connected() else "🔴 Offline"
        conn_label = tk.Label(user_frame, text=conn_text,
                             font=("Segoe UI", 9, "bold"), bg=C_HEADER, fg=conn_color)
        conn_label.pack(side=tk.LEFT, padx=(0, 8))
        
        # Botão de configuração do servidor
        _make_btn(user_frame, "🔧  Servidor", self._on_config_server, C_MUTED, size=9
                  ).pack(side=tk.LEFT, padx=(0, 8))
        
        # Mostra usuário atual
        user_label = tk.Label(user_frame, text=f"👤  {self._current_user}",
                              font=("Segoe UI", 10, "bold"), bg=C_HEADER, fg="white")
        user_label.pack(side=tk.LEFT, padx=(0, 8))
        
        # Botão de admin (só se for admin)
        if self._user_data.get("is_admin", False):
            _make_btn(user_frame, "⚙️  Admin", self._on_open_admin_panel, C_ACCENT, size=9
                      ).pack(side=tk.LEFT)

        tk.Frame(hdr, bg="#0D47A1", height=3).pack(fill=tk.X, pady=(0, 0))

    # ══════════════════════════════════════════════════════════
    #  ABA 1 — CONSULTA ONR (com scroll vertical no body)
    # ══════════════════════════════════════════════════════════
    def _build_tab1_content(self) -> None:
        """
        Constrói o conteúdo da aba 1 com Canvas+Scrollbar para scroll vertical
        quando a janela for menor que o conteúdo.
        """
        # Container com scroll
        outer = tk.Frame(self._tab1, bg=C_BG)
        outer.pack(fill=tk.BOTH, expand=True)

        canvas = tk.Canvas(outer, bg=C_BG, highlightthickness=0)
        vscroll = ttk.Scrollbar(outer, orient=tk.VERTICAL, command=canvas.yview)
        canvas.configure(yscrollcommand=vscroll.set)

        vscroll.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Frame interno que contém todo o conteúdo
        body = tk.Frame(canvas, bg=C_BG, padx=18, pady=12)
        body.columnconfigure(0, weight=1)
        body.rowconfigure(5, weight=1)

        win_id = canvas.create_window((0, 0), window=body, anchor="nw")

        def _on_body_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))

        def _on_canvas_configure(event):
            canvas.itemconfig(win_id, width=event.width)

        body.bind("<Configure>", _on_body_configure)
        canvas.bind("<Configure>", _on_canvas_configure)

        # Scroll com roda do mouse
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        # Constrói os widgets dentro do body
        self._build_folder_row(body)       # row 0
        self._build_drive_row(body)        # row 1
        self._build_protocol_box(body)     # row 2
        self._build_action_bar(body)       # row 3
        self._build_counter_bar(body)      # row 4
        self._build_results_panel(body)    # row 5

    # ── Pasta local ───────────────────────────────────────────
    def _build_folder_row(self, body: tk.Frame) -> None:
        card = _labelframe(body, " 📁  Pasta Local de Busca ", row=0)
        inner = tk.Frame(card, bg=C_CARD)
        inner.pack(fill=tk.X)
        inner.columnconfigure(0, weight=1)
        self._folder_var = tk.StringVar()
        tk.Entry(inner, textvariable=self._folder_var, font=("Segoe UI", 10),
                 state="readonly", readonlybackground="#FAFAFA",
                 relief="solid", bd=1, cursor="arrow"
                 ).grid(row=0, column=0, sticky="ew", ipady=6, padx=(0, 8))
        _make_btn(inner, "📂  Selecionar Pasta", self._on_select_folder, C_ACCENT
                  ).grid(row=0, column=1)

    # ── Pasta OneDrive ────────────────────────────────────────
    def _build_drive_row(self, body: tk.Frame) -> None:
        card = _labelframe(body, " ☁  Pasta OneDrive (fallback) ", row=1)
        inner = tk.Frame(card, bg=C_CARD)
        inner.pack(fill=tk.X)
        inner.columnconfigure(0, weight=1)
        self._drive_var = tk.StringVar(value=ONEDRIVE_DEFAULT)
        tk.Entry(inner, textvariable=self._drive_var, font=("Segoe UI", 9),
                 relief="solid", bd=1
                 ).grid(row=0, column=0, sticky="ew", ipady=5, padx=(0, 8))
        _make_btn(inner, "📂  Selecionar", self._on_select_drive, C_DRIVE, size=9
                  ).grid(row=0, column=1, padx=(0, 6))
        _make_btn(inner, "🌐  Abrir no Browser", self._on_open_drive_web, C_MUTED, size=9
                  ).grid(row=0, column=2)
        self._drive_enabled = tk.BooleanVar(value=True)
        tk.Checkbutton(card, text="Buscar no OneDrive quando não encontrado localmente",
                       variable=self._drive_enabled, font=("Segoe UI", 9),
                       bg=C_CARD, fg=C_MUTED, activebackground=C_CARD
                       ).pack(anchor=tk.W, pady=(6, 0))

    # ── Caixa de protocolos ───────────────────────────────────
    def _build_protocol_box(self, body: tk.Frame) -> None:
        card = _labelframe(body, " 📋  Protocolos para Consultar ", row=2)
        tk.Label(card, text="Cole os protocolos abaixo, um por linha  (a busca ignora maiúsculas/minúsculas):",
                 font=("Segoe UI", 9), bg=C_CARD, fg=C_MUTED).pack(anchor=tk.W, pady=(0, 5))
        frm = tk.Frame(card, bg=C_CARD)
        frm.pack(fill=tk.BOTH, expand=True)
        self._proto_txt = tk.Text(frm, height=5, font=("Courier New", 11),
                                  relief="solid", bd=1, wrap=tk.NONE, fg="gray", undo=True)
        sy = ttk.Scrollbar(frm, orient=tk.VERTICAL,   command=self._proto_txt.yview)
        sx = ttk.Scrollbar(frm, orient=tk.HORIZONTAL, command=self._proto_txt.xview)
        self._proto_txt.configure(yscrollcommand=sy.set, xscrollcommand=sx.set)
        sy.pack(side=tk.RIGHT,  fill=tk.Y)
        sx.pack(side=tk.BOTTOM, fill=tk.X)
        self._proto_txt.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self._proto_txt.insert("1.0", self._PLACEHOLDER)
        self._proto_txt.bind("<FocusIn>",  self._ph_clear)
        self._proto_txt.bind("<FocusOut>", self._ph_restore)

    # ── Barra de ações ────────────────────────────────────────
    def _build_action_bar(self, body: tk.Frame) -> None:
        bar = tk.Frame(body, bg=C_BG, pady=7)
        bar.grid(row=3, column=0, sticky="ew")
        self._btn_consult = _make_btn(bar, "🔍   CONSULTAR", self._on_consult, C_ACCENT, size=11)
        self._btn_consult.pack(side=tk.LEFT, padx=(0, 8))
        _make_btn(bar, "🗑   LIMPAR", self._on_clear, C_MUTED, size=11).pack(side=tk.LEFT, padx=(0, 8))
        self._btn_open = _make_btn(bar, "📂   ABRIR", self._on_open_file, C_SUCCESS, size=11)
        self._btn_open.pack(side=tk.LEFT, padx=(0, 8))
        self._btn_open.config(state="disabled")
        self._btn_print = _make_btn(bar, "🖨   IMPRIMIR", self._on_print_file, C_PRINT, size=11)
        self._btn_print.pack(side=tk.LEFT, padx=(0, 8))
        self._btn_print.config(state="disabled")
        self._btn_download = _make_btn(bar, "⬇   BAIXAR DO DRIVE", self._on_download_from_drive, C_DRIVE, size=11)
        self._btn_download.pack(side=tk.LEFT)
        self._btn_download.config(state="disabled")

    # ── Barra de contadores ───────────────────────────────────
    def _build_counter_bar(self, body: tk.Frame) -> None:
        bar = tk.Frame(body, bg=C_COUNTER_BG, relief="solid", bd=1)
        bar.grid(row=4, column=0, sticky="ew", pady=(0, 6))
        inner = tk.Frame(bar, bg=C_COUNTER_BG, pady=7)
        inner.pack()
        lbl_kw = dict(font=("Segoe UI", 10, "bold"), bg=C_COUNTER_BG)
        sep_kw = dict(text="│", bg=C_COUNTER_BG, fg="#B0BEC5", font=("Segoe UI", 14))
        self._lbl_total    = tk.Label(inner, text="Total: 0",              fg="#1A3C5E", **lbl_kw)
        self._lbl_total.pack(side=tk.LEFT, padx=14)
        tk.Label(inner, **sep_kw).pack(side=tk.LEFT)
        self._lbl_found    = tk.Label(inner, text="✅  Local: 0",          fg=C_SUCCESS, **lbl_kw)
        self._lbl_found.pack(side=tk.LEFT, padx=14)
        tk.Label(inner, **sep_kw).pack(side=tk.LEFT)
        self._lbl_drive    = tk.Label(inner, text="☁  Drive: 0",           fg=C_DRIVE,   **lbl_kw)
        self._lbl_drive.pack(side=tk.LEFT, padx=14)
        tk.Label(inner, **sep_kw).pack(side=tk.LEFT)
        self._lbl_notfound = tk.Label(inner, text="❌  Não encontrados: 0", fg=C_DANGER,  **lbl_kw)
        self._lbl_notfound.pack(side=tk.LEFT, padx=14)

    # ── Painel de resultados ──────────────────────────────────
    def _build_results_panel(self, body: tk.Frame) -> None:
        card = _labelframe(body, " 📊  Resultados ", row=5, expand=True)
        frm = tk.Frame(card, bg=C_CARD)
        frm.pack(fill=tk.BOTH, expand=True)
        cols = ("status", "protocolo", "arquivo", "caminho")
        self._tree = ttk.Treeview(frm, columns=cols, show="headings", selectmode="extended")
        self._tree.heading("status",    text="Status",             anchor=tk.CENTER)
        self._tree.heading("protocolo", text="Protocolo",          anchor=tk.W)
        self._tree.heading("arquivo",   text="Arquivo encontrado", anchor=tk.W)
        self._tree.heading("caminho",   text="Caminho completo",   anchor=tk.W)
        self._tree.column("status",    width=185, minwidth=150, anchor=tk.CENTER, stretch=False)
        self._tree.column("protocolo", width=160, minwidth=130, stretch=False)
        self._tree.column("arquivo",   width=280, minwidth=180)
        self._tree.column("caminho",   width=360, minwidth=200)
        self._tree.tag_configure("found",    foreground=C_SUCCESS)
        self._tree.tag_configure("drive",    foreground=C_DRIVE)
        self._tree.tag_configure("notfound", foreground=C_DANGER)
        sy = ttk.Scrollbar(frm, orient=tk.VERTICAL,   command=self._tree.yview)
        sx = ttk.Scrollbar(frm, orient=tk.HORIZONTAL, command=self._tree.xview)
        self._tree.configure(yscrollcommand=sy.set, xscrollcommand=sx.set)
        sy.pack(side=tk.RIGHT,  fill=tk.Y)
        sx.pack(side=tk.BOTTOM, fill=tk.X)
        self._tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self._tree.bind("<<TreeviewSelect>>", self._on_tree_select)
        self._tree.bind("<Double-1>",         lambda _e: self._on_open_file())

    # ══════════════════════════════════════════════════════════
    #  ABA 2 — PLANILHA DE CONTROLE
    # ══════════════════════════════════════════════════════════
    def _build_tab2_content(self) -> None:
        """Instancia o PlanilhaTab dentro da aba 2."""
        self._planilha_tab = PlanilhaTab(self._tab2_outer, self._status_var)
        self._planilha_tab.pack(fill=tk.BOTH, expand=True, padx=14, pady=10)

    # ══════════════════════════════════════════════════════════
    #  LÓGICA DE NEGÓCIO — ABA 1 (idêntica ao ConsultaCertidoesApp)
    # ══════════════════════════════════════════════════════════

    # ── Placeholder ───────────────────────────────────────────
    def _ph_clear(self, _event=None) -> None:
        if self._ph_active:
            self._proto_txt.delete("1.0", tk.END)
            self._proto_txt.config(fg="#212121")
            self._ph_active = False

    def _ph_restore(self, _event=None) -> None:
        if not self._proto_txt.get("1.0", tk.END).strip():
            self._proto_txt.insert("1.0", self._PLACEHOLDER)
            self._proto_txt.config(fg="gray")
            self._ph_active = True

    # ── Ações ─────────────────────────────────────────────────
    def _on_select_folder(self) -> None:
        path = filedialog.askdirectory(title="Selecionar pasta local de certidões")
        if path:
            self._folder_var.set(path)
            self._status_var.set(f"Pasta local selecionada: {path}")

    def _on_select_drive(self) -> None:
        path = filedialog.askdirectory(title="Selecionar pasta do OneDrive")
        if path:
            self._drive_var.set(path)
            self._status_var.set(f"Pasta OneDrive selecionada: {path}")

    def _on_open_drive_web(self) -> None:
        try:
            import webbrowser
            webbrowser.open(ONEDRIVE_WEB_URL)
        except Exception as exc:
            messagebox.showerror("Erro", str(exc))

    def _on_consult(self) -> None:
        folder     = self._folder_var.get().strip()
        drive_path = self._drive_var.get().strip()
        use_drive  = self._drive_enabled.get()
        protocols  = self._get_protocols()

        if not folder:
            messagebox.showwarning("Atenção", "Selecione a pasta local antes de consultar.")
            return
        if not os.path.isdir(folder):
            messagebox.showerror("Erro", f"Pasta local não encontrada:\n{folder}")
            return
        if not protocols:
            messagebox.showwarning("Atenção", "Informe ao menos um protocolo.")
            return

        if use_drive and drive_path and not os.path.isdir(drive_path):
            resp = messagebox.askyesno(
                "Pasta OneDrive não encontrada",
                f"A pasta OneDrive não foi localizada:\n{drive_path}\n\n"
                "Continuar apenas com a busca local?",
            )
            if not resp:
                return
            use_drive = False

        self._tree.delete(*self._tree.get_children())
        self._found_files.clear()
        self._drive_files.clear()
        self._btn_open.config(state="disabled")
        self._btn_print.config(state="disabled")
        self._btn_download.config(state="disabled")
        self._update_counters(0, 0, 0, 0)

        self._btn_consult.config(state="disabled", text="⏳  Consultando...")
        self._status_var.set("Indexando arquivos … aguarde")
        self.update_idletasks()

        threading.Thread(
            target=self._search_worker,
            args=(folder, drive_path if use_drive else "", protocols),
            daemon=True,
        ).start()

    def _search_worker(self, local_folder: str, drive_folder: str, protocols: list) -> None:
        def index_folder(folder: str) -> list[tuple[str, str]]:
            files = []
            for dirpath, _dirs, fnames in os.walk(folder):
                for fname in fnames:
                    files.append((fname, os.path.join(dirpath, fname)))
            return files

        local_files = index_folder(local_folder)
        drive_files = index_folder(drive_folder) if drive_folder else []

        results = []
        local_count = drive_count = not_found_count = 0

        for proto in protocols:
            proto_lower = proto.lower()
            local_matches = [(f, p) for f, p in local_files if proto_lower in f.lower()]
            if local_matches:
                local_count += 1
                for fname, fpath in local_matches:
                    results.append(("found", proto, fname, fpath))
                continue
            drive_matches = [(f, p) for f, p in drive_files if proto_lower in f.lower()]
            if drive_matches:
                drive_count += 1
                for fname, fpath in drive_matches:
                    results.append(("drive", proto, fname, fpath))
            else:
                not_found_count += 1
                results.append(("notfound", proto, "", ""))

        self.after(0, self._show_results, results, len(protocols),
                   local_count, drive_count, not_found_count)

    def _show_results(self, results, total, local, drive, not_found) -> None:
        for status, proto, fname, fpath in results:
            if status == "found":
                iid = self._tree.insert("", tk.END,
                                        values=("✅  Local", proto, fname, fpath),
                                        tags=("found",))
                self._found_files[iid] = fpath
            elif status == "drive":
                iid = self._tree.insert("", tk.END,
                                        values=("☁  OneDrive", proto, fname, fpath),
                                        tags=("drive",))
                self._found_files[iid] = fpath
                self._drive_files.add(iid)
            else:
                self._tree.insert("", tk.END,
                                  values=("❌  Não encontrado", proto, "—", "—"),
                                  tags=("notfound",))
        self._update_counters(total, local, drive, not_found)
        self._btn_consult.config(state="normal", text="🔍   CONSULTAR")
        self._status_var.set(
            f"Consulta finalizada  •  {total} protocolo(s)  •  "
            f"{local} local / {drive} OneDrive / {not_found} não encontrado(s)"
        )

    def _on_tree_select(self, _event=None) -> None:
        sel = self._tree.selection()
        found_sel = [iid for iid in sel if iid in self._found_files]
        drive_sel = [iid for iid in found_sel if iid in self._drive_files]
        self._btn_open.config(state="normal" if found_sel else "disabled")
        if found_sel:
            n = len(found_sel)
            self._btn_print.config(state="normal",
                                   text=f"🖨  IMPRIMIR ({n})" if n > 1 else "🖨   IMPRIMIR")
        else:
            self._btn_print.config(state="disabled", text="🖨   IMPRIMIR")
        if drive_sel:
            n = len(drive_sel)
            self._btn_download.config(state="normal",
                                      text=f"⬇   BAIXAR DO DRIVE ({n})" if n > 1 else "⬇   BAIXAR DO DRIVE")
        else:
            self._btn_download.config(state="disabled", text="⬇   BAIXAR DO DRIVE")

    def _on_open_file(self) -> None:
        sel = self._tree.selection()
        if not sel:
            return
        iid = sel[0]
        if iid not in self._found_files:
            messagebox.showinfo("Aviso", "Selecione um item ✅/☁ para abrir.")
            return
        path = self._found_files[iid]
        if not os.path.exists(path):
            messagebox.showerror("Arquivo não encontrado", f"Não localizado:\n{path}")
            return
        try:
            os.startfile(path)
        except Exception as exc:
            messagebox.showerror("Erro ao abrir arquivo", str(exc))

    def _open_from_zip(self, zip_path: str) -> None:
        """Extrai PDFs de um ZIP e abre o PDF extraído. Se múltiplos, mostra diálogo."""
        try:
            with zipfile.ZipFile(zip_path, "r") as zf:
                pdf_names = [n for n in zf.namelist()
                             if n.lower().endswith(".pdf") and not n.startswith("__MACOSX")]
        except zipfile.BadZipFile:
            messagebox.showerror("ZIP inválido", f"Não é um ZIP válido:\n{zip_path}")
            return
        if not pdf_names:
            messagebox.showwarning("Nenhum PDF", f"ZIP sem PDFs:\n{zip_path}")
            return
        tmp_dir = tempfile.mkdtemp(prefix="onr_open_")
        self._temp_dirs.append(tmp_dir)
        try:
            with zipfile.ZipFile(zip_path, "r") as zf:
                for name in pdf_names:
                    zf.extract(name, tmp_dir)
        except Exception as exc:
            messagebox.showerror("Erro ao extrair ZIP", str(exc))
            return
        if len(pdf_names) == 1:
            try:
                os.startfile(os.path.join(tmp_dir, pdf_names[0]))
            except Exception as exc:
                messagebox.showerror("Erro ao abrir PDF", str(exc))
            return
        self._show_zip_open_dialog(zip_path, tmp_dir, pdf_names)

    def _show_zip_open_dialog(self, zip_path: str, tmp_dir: str, pdf_names: list) -> None:
        """Diálogo para escolher qual PDF do ZIP abrir."""
        dialog = tk.Toplevel(self)
        dialog.title("Selecionar PDF para Abrir")
        dialog.geometry("560x400")
        dialog.resizable(True, True)
        dialog.configure(bg=C_BG)
        dialog.grab_set()
        dialog.update_idletasks()
        px = self.winfo_x() + (self.winfo_width()  - 560) // 2
        py = self.winfo_y() + (self.winfo_height() - 400) // 2
        dialog.geometry(f"560x400+{px}+{py}")
        hdr = tk.Frame(dialog, bg=C_HEADER)
        hdr.pack(fill=tk.X)
        tk.Label(hdr, text="📂  Abrir PDF do ZIP",
                 font=("Segoe UI", 13, "bold"), bg=C_HEADER, fg="white", pady=10).pack()
        tk.Label(hdr, text=f"ZIP: {os.path.basename(zip_path)}",
                 font=("Segoe UI", 8), bg=C_HEADER, fg="#90CAF9").pack(pady=(0, 4))
        tk.Frame(hdr, bg="#0D47A1", height=2).pack(fill=tk.X)
        tk.Label(dialog, text="Selecione o PDF que deseja abrir:",
                 font=("Segoe UI", 9), bg=C_BG, fg=C_MUTED, anchor=tk.W
                 ).pack(fill=tk.X, padx=14, pady=(10, 4))
        list_frame = tk.Frame(dialog, bg=C_CARD, relief="solid", bd=1)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=14, pady=(0, 8))
        lb_scroll = ttk.Scrollbar(list_frame, orient=tk.VERTICAL)
        listbox = tk.Listbox(list_frame, font=("Segoe UI", 10),
                             yscrollcommand=lb_scroll.set,
                             selectmode=tk.SINGLE,
                             bg=C_CARD, fg="#212121",
                             selectbackground="#BBDEFB", selectforeground="#0D47A1",
                             relief="flat", bd=0)
        lb_scroll.config(command=listbox.yview)
        lb_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=4, pady=4)
        for name in pdf_names:
            listbox.insert(tk.END, f"  {os.path.basename(name)}")
        if pdf_names:
            listbox.selection_set(0)
        btn_bar = tk.Frame(dialog, bg=C_BG)
        btn_bar.pack(fill=tk.X, padx=14, pady=(0, 12))

        def do_open():
            sel_idx = listbox.curselection()
            if not sel_idx:
                messagebox.showwarning("Nenhum selecionado", "Selecione um PDF.", parent=dialog)
                return
            chosen = pdf_names[sel_idx[0]]
            dialog.destroy()
            try:
                os.startfile(os.path.join(tmp_dir, chosen))
            except Exception as exc:
                messagebox.showerror("Erro ao abrir PDF", str(exc))

        _make_btn(btn_bar, "📂  Abrir selecionado", do_open, C_SUCCESS).pack(side=tk.LEFT, padx=(0, 8))
        _make_btn(btn_bar, "Cancelar", dialog.destroy, C_MUTED, size=9).pack(side=tk.RIGHT)
        listbox.bind("<Double-1>", lambda _e: do_open())

    def _on_download_from_drive(self) -> None:
        sel = self._tree.selection()
        drive_sel = [iid for iid in sel if iid in self._drive_files]
        if not drive_sel:
            return
        local_folder = self._folder_var.get().strip()
        if not local_folder or not os.path.isdir(local_folder):
            messagebox.showerror("Pasta local não definida",
                                 "Selecione uma pasta local válida antes de baixar.")
            return
        items = [(iid, self._found_files[iid]) for iid in drive_sel]
        nomes = "\n".join(f"  • {os.path.basename(p)}" for _, p in items)
        ok = messagebox.askyesno("Confirmar download",
                                 f"Copiar {len(items)} arquivo(s) do OneDrive para:\n"
                                 f"{local_folder}\n\n{nomes}")
        if not ok:
            return
        erros, copiados = [], []
        for iid, src_path in items:
            fname = os.path.basename(src_path)
            dst_path = os.path.join(local_folder, fname)
            if os.path.exists(dst_path):
                resp = messagebox.askyesno("Arquivo já existe",
                                           f"O arquivo já existe na pasta local:\n{fname}\n\nSobrescrever?")
                if not resp:
                    continue
            try:
                shutil.copy2(src_path, dst_path)
                copiados.append((iid, fname, dst_path))
            except Exception as exc:
                erros.append(f"{fname}: {exc}")
        for iid, fname, dst_path in copiados:
            self._tree.item(iid,
                            values=("✅  Local", self._tree.item(iid, "values")[1], fname, dst_path),
                            tags=("found",))
            self._found_files[iid] = dst_path
            self._drive_files.discard(iid)
        if erros:
            messagebox.showerror("Erros ao copiar",
                                 "Alguns arquivos não puderam ser copiados:\n\n" + "\n".join(erros))
        elif copiados:
            messagebox.showinfo("Download concluído",
                                f"{len(copiados)} arquivo(s) copiado(s) para:\n{local_folder}")
            self._status_var.set(
                f"☁→📁  {len(copiados)} arquivo(s) baixado(s) do OneDrive para a pasta local.")
        self._recount()

    def _on_clear(self) -> None:
        self._ph_clear()
        self._proto_txt.delete("1.0", tk.END)
        self._ph_restore()
        self._tree.delete(*self._tree.get_children())
        self._found_files.clear()
        self._drive_files.clear()
        self._btn_open.config(state="disabled")
        self._btn_print.config(state="disabled")
        self._btn_download.config(state="disabled", text="⬇   BAIXAR DO DRIVE")
        self._update_counters(0, 0, 0, 0)
        self._status_var.set("Campos limpos. Pronto para nova consulta.")

    # ── Impressão ─────────────────────────────────────────────
    def _on_print_file(self) -> None:
        sel = self._tree.selection()
        if not sel:
            return
        found_sel = [iid for iid in sel if iid in self._found_files]
        if not found_sel:
            messagebox.showinfo("Aviso", "Selecione ao menos um item ✅/☁ para imprimir.")
            return
        paths, missing = [], []
        for iid in found_sel:
            p = self._found_files[iid]
            (paths if os.path.exists(p) else missing).append(p)
        if missing:
            messagebox.showwarning("Arquivos não encontrados",
                                   "Ignorados (não localizados no disco):\n" + "\n".join(missing))
        if not paths:
            return
        if len(paths) > 1:
            nomes = "\n".join(f"  • {os.path.basename(p)}" for p in paths)
            if not messagebox.askyesno("Confirmar impressão",
                                       f"Imprimir {len(paths)} arquivo(s)?\n\n{nomes}"):
                return
        for path in paths:
            if os.path.splitext(path)[1].lower() == ".zip":
                self._print_from_zip(path)
            else:
                self._print_file_direct(path)

    def _print_file_direct(self, path: str) -> None:
        fname = os.path.basename(path)
        ext   = os.path.splitext(path)[1].lower()
        if ext == ".pdf":
            ok, method = self._try_print_pdf(path)
        else:
            ok, method = self._shellexec_print(path), "ShellExecute"
        if ok:
            self._status_var.set(f"🖨  Enviado para impressão via {method}: {fname}")
        else:
            if messagebox.askyesno("Impressão automática indisponível",
                                   f"Não foi possível enviar diretamente.\n\nAbrir para imprimir manualmente?\n{path}"):
                try:
                    os.startfile(path)
                except Exception as exc:
                    messagebox.showerror("Erro ao abrir", str(exc))

    def _try_print_pdf(self, path: str) -> tuple[bool, str]:
        sumatra = self._find_exe([
            r"C:\Program Files\SumatraPDF\SumatraPDF.exe",
            r"C:\Program Files (x86)\SumatraPDF\SumatraPDF.exe",
        ], which="SumatraPDF")
        if sumatra:
            try:
                subprocess.Popen([sumatra, "-print-to-default", "-silent", path],
                                 stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                return True, "SumatraPDF"
            except Exception:
                pass
        acrobat = self._find_exe([
            r"C:\Program Files\Adobe\Acrobat DC\Acrobat\Acrobat.exe",
            r"C:\Program Files (x86)\Adobe\Acrobat DC\Acrobat\Acrobat.exe",
            r"C:\Program Files\Adobe\Acrobat Reader DC\Reader\AcroRd32.exe",
            r"C:\Program Files (x86)\Adobe\Acrobat Reader DC\Reader\AcroRd32.exe",
        ], which="AcroRd32")
        if acrobat:
            try:
                subprocess.Popen([acrobat, "/p", "/h", path],
                                 stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                return True, "Adobe Reader"
            except Exception:
                pass
        foxit = self._find_exe([
            r"C:\Program Files\Foxit Software\Foxit PDF Reader\FoxitPDFReader.exe",
            r"C:\Program Files (x86)\Foxit Software\Foxit Reader\FoxitReader.exe",
        ], which="FoxitReader")
        if foxit:
            try:
                subprocess.Popen([foxit, "/p", path],
                                 stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                return True, "Foxit Reader"
            except Exception:
                pass
        default_printer = self._get_default_printer()
        if default_printer:
            try:
                import ctypes
                ret = ctypes.windll.shell32.ShellExecuteW(
                    None, "printto", path, f'"{default_printer}"', None, 0)
                if ret > 32:
                    return True, f"impressora padrão ({default_printer})"
            except Exception:
                pass
        if self._shellexec_print(path):
            return True, "ShellExecute"
        return False, ""

    @staticmethod
    def _find_exe(paths: list, which: str | None = None) -> str | None:
        for p in paths:
            if os.path.isfile(p):
                return p
        if which:
            found = shutil.which(which)
            if found:
                return found
        return None

    @staticmethod
    def _get_default_printer() -> str | None:
        try:
            import win32print
            return win32print.GetDefaultPrinter()
        except ImportError:
            pass
        try:
            import winreg
            key = winreg.OpenKey(winreg.HKEY_CURRENT_USER,
                                 r"Software\Microsoft\Windows NT\CurrentVersion\Windows")
            printer, _ = winreg.QueryValueEx(key, "Device")
            winreg.CloseKey(key)
            return printer.split(",")[0].strip()
        except Exception:
            return None

    @staticmethod
    def _shellexec_print(path: str) -> bool:
        try:
            import ctypes
            ret = ctypes.windll.shell32.ShellExecuteW(None, "print", path, None, None, 0)
            return ret > 32
        except Exception:
            return False

    def _print_from_zip(self, zip_path: str) -> None:
        try:
            with zipfile.ZipFile(zip_path, "r") as zf:
                pdf_names = [n for n in zf.namelist()
                             if n.lower().endswith(".pdf") and not n.startswith("__MACOSX")]
        except zipfile.BadZipFile:
            messagebox.showerror("ZIP inválido", f"Não é um ZIP válido:\n{zip_path}")
            return
        if not pdf_names:
            messagebox.showwarning("Nenhum PDF", f"ZIP sem PDFs:\n{zip_path}")
            return
        tmp_dir = tempfile.mkdtemp(prefix="onr_print_")
        self._temp_dirs.append(tmp_dir)
        try:
            with zipfile.ZipFile(zip_path, "r") as zf:
                for name in pdf_names:
                    zf.extract(name, tmp_dir)
        except Exception as exc:
            messagebox.showerror("Erro ao extrair ZIP", str(exc))
            return
        if len(pdf_names) == 1:
            self._print_file_direct(os.path.join(tmp_dir, pdf_names[0]))
            return
        self._show_zip_print_dialog(zip_path, tmp_dir, pdf_names)

    def _show_zip_print_dialog(self, zip_path: str, tmp_dir: str, pdf_names: list) -> None:
        dialog = tk.Toplevel(self)
        dialog.title("Selecionar PDFs para Imprimir")
        dialog.geometry("560x400")
        dialog.resizable(True, True)
        dialog.configure(bg=C_BG)
        dialog.grab_set()
        dialog.update_idletasks()
        px = self.winfo_x() + (self.winfo_width()  - 560) // 2
        py = self.winfo_y() + (self.winfo_height() - 400) // 2
        dialog.geometry(f"560x400+{px}+{py}")
        hdr = tk.Frame(dialog, bg=C_HEADER)
        hdr.pack(fill=tk.X)
        tk.Label(hdr, text="🖨  Imprimir PDFs do ZIP",
                 font=("Segoe UI", 13, "bold"), bg=C_HEADER, fg="white", pady=10).pack()
        tk.Label(hdr, text=f"ZIP: {os.path.basename(zip_path)}",
                 font=("Segoe UI", 8), bg=C_HEADER, fg="#90CAF9").pack(pady=(0, 4))
        tk.Frame(hdr, bg="#0D47A1", height=2).pack(fill=tk.X)
        tk.Label(dialog, text="Marque os PDFs que deseja imprimir:",
                 font=("Segoe UI", 9), bg=C_BG, fg=C_MUTED, anchor=tk.W
                 ).pack(fill=tk.X, padx=14, pady=(10, 4))
        list_frame = tk.Frame(dialog, bg=C_CARD, relief="solid", bd=1)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=14, pady=(0, 8))
        canvas = tk.Canvas(list_frame, bg=C_CARD, highlightthickness=0)
        scroll = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=canvas.yview)
        inner  = tk.Frame(canvas, bg=C_CARD)
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=inner, anchor="nw")
        canvas.configure(yscrollcommand=scroll.set)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        check_vars: list[tk.BooleanVar] = []
        for i, name in enumerate(pdf_names):
            var = tk.BooleanVar(value=True)
            check_vars.append(var)
            row_bg = C_CARD if i % 2 == 0 else "#F7F9FC"
            tk.Checkbutton(inner, text=f"  {os.path.basename(name)}", variable=var,
                           font=("Segoe UI", 10), bg=row_bg, fg="#212121",
                           activebackground=row_bg, anchor=tk.W,
                           ).pack(fill=tk.X, padx=8, pady=3, ipadx=4, ipady=3)
        btn_bar = tk.Frame(dialog, bg=C_BG)
        btn_bar.pack(fill=tk.X, padx=14, pady=(0, 12))

        def do_print():
            selected = [os.path.join(tmp_dir, pdf_names[i])
                        for i, v in enumerate(check_vars) if v.get()]
            if not selected:
                messagebox.showwarning("Nenhum selecionado", "Marque ao menos um PDF.", parent=dialog)
                return
            dialog.destroy()
            for p in selected:
                self._print_file_direct(p)

        _make_btn(btn_bar, "🖨  Imprimir selecionados", do_print, C_PRINT).pack(side=tk.LEFT, padx=(0, 8))
        _make_btn(btn_bar, "✔  Todos",  lambda: [v.set(True)  for v in check_vars], C_ACCENT, size=9).pack(side=tk.LEFT, padx=(0, 4))
        _make_btn(btn_bar, "✖  Nenhum", lambda: [v.set(False) for v in check_vars], C_MUTED,  size=9).pack(side=tk.LEFT)
        _make_btn(btn_bar, "Cancelar",  dialog.destroy, C_MUTED, size=9).pack(side=tk.RIGHT)

    # ── Utilitários ───────────────────────────────────────────
    def _get_protocols(self) -> list:
        if self._ph_active:
            return []
        return [line.strip()
                for line in self._proto_txt.get("1.0", tk.END).splitlines()
                if line.strip()]

    def _update_counters(self, total, local, drive, not_found) -> None:
        self._lbl_total.config(text=f"Total: {total}")
        self._lbl_found.config(text=f"✅  Local: {local}")
        self._lbl_drive.config(text=f"☁  Drive: {drive}")
        self._lbl_notfound.config(text=f"❌  Não encontrados: {not_found}")

    def _recount(self) -> None:
        local = drive = notfound = total = 0
        for iid in self._tree.get_children():
            total += 1
            tags = self._tree.item(iid, "tags")
            if "found"    in tags: local    += 1
            elif "drive"  in tags: drive    += 1
            elif "notfound" in tags: notfound += 1
        self._update_counters(total, local, drive, notfound)

    def _center_window(self) -> None:
        self.update_idletasks()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        w,  h  = self.winfo_width(), self.winfo_height()
        self.geometry(f"{w}x{h}+{(sw - w) // 2}+{(sh - h) // 2}")

    def _on_open_admin_panel(self) -> None:
        """Abre o painel de administração."""
        if not self._user_data.get("is_admin", False):
            messagebox.showerror("Acesso Negado", "Você não tem permissão para acessar o painel de administração.")
            return
        AdminPanel(self, self._user_manager, self._current_user)

    def _on_config_server(self) -> None:
        """Abre diálogo para configurar URL do servidor."""
        dialog = tk.Toplevel(self)
        dialog.title("Configurar Servidor")
        dialog.geometry("500x300")
        dialog.resizable(False, False)
        dialog.configure(bg=C_BG)
        dialog.grab_set()
        self._center_dialog(dialog)

        # Header
        hdr = tk.Frame(dialog, bg=C_HEADER)
        hdr.pack(fill=tk.X)
        tk.Label(hdr, text="🔧  Configuração do Servidor",
                 font=("Segoe UI", 13, "bold"), bg=C_HEADER, fg="white", pady=10).pack()
        tk.Frame(hdr, bg="#0D47A1", height=2).pack(fill=tk.X)

        # Container
        container = tk.Frame(dialog, bg=C_BG, padx=25, pady=20)
        container.pack(fill=tk.BOTH, expand=True)

        # Campo URL do servidor
        tk.Label(container, text="URL do Servidor:", font=("Segoe UI", 10, "bold"),
                 bg=C_BG, fg="#333").pack(anchor=tk.W, pady=(0, 5))
        url_var = tk.StringVar(value=self._user_manager.get_server_url())
        url_entry = tk.Entry(container, textvariable=url_var, font=("Segoe UI", 11),
                            relief="solid", bd=1, width=40)
        url_entry.pack(ipady=8, pady=(0, 15))

        # Status atual
        is_connected = self._user_manager.is_connected()
        status_color = C_SUCCESS if is_connected else C_DANGER
        status_text = "🟢 Conectado" if is_connected else "🔴 Desconectado"
        tk.Label(container, text=f"Status: {status_text}",
                 font=("Segoe UI", 10, "bold"), bg=C_BG, fg=status_color).pack(anchor=tk.W, pady=(0, 20))

        tk.Label(container, text="Exemplos:",
                 font=("Segoe UI", 9, "bold"), bg=C_BG, fg="#666").pack(anchor=tk.W)
        tk.Label(container, text="• https://brejeiro.onrender.com (servidor cloud)",
                 font=("Segoe UI", 9), bg=C_BG, fg="#666").pack(anchor=tk.W)
        tk.Label(container, text="• http://localhost:5000 (local)",
                 font=("Segoe UI", 9), bg=C_BG, fg="#666").pack(anchor=tk.W)
        tk.Label(container, text="• http://192.168.1.100:5000 (rede local)",
                 font=("Segoe UI", 9), bg=C_BG, fg="#666").pack(anchor=tk.W)

        # Botões
        btn_frame = tk.Frame(container, bg=C_BG)
        btn_frame.pack(fill=tk.X, pady=(20, 0))

        def on_save():
            new_url = url_var.get().strip()
            if not new_url:
                messagebox.showwarning("Campo obrigatório",
                                       "A URL do servidor é obrigatória.", parent=dialog)
                return
            
            if self._user_manager.set_server_url(new_url):
                messagebox.showinfo("Sucesso", "Configuração salva com sucesso.", parent=dialog)
                dialog.destroy()
                # Recarrega a aplicação para aplicar mudanças
                self._on_close()
            else:
                messagebox.showerror("Erro", "Erro ao salvar configuração.", parent=dialog)

        def on_test():
            new_url = url_var.get().strip()
            if not new_url:
                messagebox.showwarning("Campo obrigatório",
                                       "A URL do servidor é obrigatória.", parent=dialog)
                return
            
            # Testa conexão temporária
            temp_manager = UserManager(new_url)
            if temp_manager.is_connected():
                messagebox.showinfo("Teste de Conexão", "Conexão estabelecida com sucesso!", parent=dialog)
            else:
                messagebox.showerror("Teste de Conexão", "Não foi possível conectar ao servidor.", parent=dialog)

        _make_btn(btn_frame, "💾  Salvar", on_save, C_SUCCESS, size=10
                  ).pack(side=tk.LEFT, expand=True, fill=tk.X, padx=(0, 8))
        _make_btn(btn_frame, "🧪  Testar", on_test, C_ACCENT, size=10
                  ).pack(side=tk.LEFT, expand=True, fill=tk.X, padx=(0, 8))
        _make_btn(btn_frame, "Cancelar", dialog.destroy, C_MUTED, size=10
                  ).pack(side=tk.LEFT, expand=True, fill=tk.X)

        url_entry.focus_set()

    def _center_dialog(self, dialog: tk.Toplevel) -> None:
        dialog.update_idletasks()
        px = self.winfo_x() + (self.winfo_width() - 500) // 2
        py = self.winfo_y() + (self.winfo_height() - 300) // 2
        dialog.geometry(f"500x300+{px}+{py}")

    def _on_close(self) -> None:
        for tmp in self._temp_dirs:
            shutil.rmtree(tmp, ignore_errors=True)
        self.destroy()


# ══════════════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════════════
if __name__ == "__main__":
    # Inicializa gerenciador de usuários
    user_manager = UserManager()
    
    # Cria app temporário para verificação de atualizações
    temp_app = tk.Tk()
    temp_app.withdraw()
    
    # Verifica atualizações antes do login
    has_update, local_ver, remote_ver = UpdateManager.has_update()
    if has_update:
        resp = messagebox.askyesno(
            "Atualização Disponível",
            f"Há uma nova versão disponível!\n\n"
            f"Versão atual: {local_ver}\n"
            f"Nova versão: {remote_ver}\n\n"
            f"Deseja atualizar agora?"
        )
        if resp:
            # Baixa e aplica atualização
            new_exe = UpdateManager.download_update()
            if new_exe:
                if UpdateManager.apply_update(new_exe):
                    messagebox.showinfo(
                        "Atualização Concluída",
                        "O aplicativo foi atualizado com sucesso.\n"
                        "Ele será reiniciado agora."
                    )
                    # Reinicia o aplicativo
                    os.execv(sys.executable, [sys.executable] + sys.argv)
                else:
                    messagebox.showerror(
                        "Erro na Atualização",
                        "Não foi possível aplicar a atualização.\n"
                        "Entre em contato com o suporte."
                    )
            else:
                messagebox.showerror(
                    "Erro no Download",
                    "Não foi possível baixar a atualização.\n"
                    "Verifique sua conexão ou a pasta de atualizações."
                )
    
    # Mostra tela de login
    login = LoginDialog(temp_app, user_manager)
    login.wait_window(login)
    
    temp_app.destroy()
    
    if login.is_authenticated():
        current_user, user_data = login.get_authenticated_user()
        app = MainApp(user_manager, current_user, user_data)
        app.mainloop()
    else:
        # Login foi cancelado
        pass
